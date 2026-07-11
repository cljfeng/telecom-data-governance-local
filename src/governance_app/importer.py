import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from governance_app.config import AppConfig
from governance_app.db import connect
from governance_app.models import LedgerType, ValidationErrorDetail
from governance_app.recent_files import record_recent_file
from governance_app.templates import (
    EXPECTED_SHEETS,
    HEADER_ROWS,
    canonical_header,
    required_headers_for,
    workbook_sheet_for,
)
from governance_app.workflow import _new_batch_code, transition_batch_in_conn


@dataclass(frozen=True)
class ImportResult:
    batch_id: int | None
    errors: list[ValidationErrorDetail] = field(default_factory=list)
    ledger_counts: dict[str, int] = field(default_factory=dict)


def import_workbook(config: AppConfig, workbook_path: Path, strategy: str = "new", batch_id: int | None = None) -> ImportResult:
    started_at = perf_counter()
    wb = load_workbook(workbook_path, data_only=True)
    errors: list[ValidationErrorDetail] = []
    parsed: dict[LedgerType, tuple[str, list[tuple[int, dict[str, Any]]]]] = {}

    for canonical_sheet_name, ledger_type in EXPECTED_SHEETS.items():
        sheet_name = workbook_sheet_for(wb.sheetnames, ledger_type)
        if sheet_name is None:
            errors.append(ValidationErrorDetail(0, canonical_sheet_name, "缺少必需 sheet"))
            continue
        ws = wb[sheet_name]
        headers = _headers(ws, HEADER_ROWS[ledger_type], ledger_type)
        missing = [name for name in required_headers_for(ledger_type) if name not in headers]
        for name in missing:
            errors.append(ValidationErrorDetail(1, name, "缺少必需字段"))
        if missing:
            continue
        parsed[ledger_type] = (sheet_name, _data_rows(ws, headers, HEADER_ROWS[ledger_type] + 1))

    if errors:
        return ImportResult(batch_id=None, errors=errors)

    if strategy not in {"new", "append", "replace"}:
        raise ValueError("invalid import strategy")
    if strategy in {"append", "replace"} and batch_id is None:
        raise ValueError("batch_id is required")

    with connect(config) as conn:
        if strategy == "new":
            batch_name = _clean_batch_name(workbook_path.stem)
            batch_id = conn.execute(
                "insert into import_batches(source_file, name, batch_code, status) values (?, ?, ?, ?)",
                (str(workbook_path), batch_name, _new_batch_code(), "imported"),
            ).lastrowid
            operation = "import"
            message = f"导入台账：{workbook_path.name}"
        else:
            batch = conn.execute("select id, is_archived from import_batches where id = ?", (batch_id,)).fetchone()
            if batch is None:
                raise ValueError("batch not found")
            if batch["is_archived"]:
                raise ValueError("batch is archived")
            if strategy == "replace":
                _clear_batch_data(conn, batch_id)
                operation = "import_replace"
                message = f"覆盖导入台账：{workbook_path.name}"
            else:
                operation = "import_append"
                message = f"追加导入台账：{workbook_path.name}"
            conn.execute(
                "update import_batches set source_file = ?, name = coalesce(name, ?) where id = ?",
                (str(workbook_path), _clean_batch_name(workbook_path.stem), batch_id),
            )
            transition_batch_in_conn(conn, batch_id, "import")
        conn.execute(
            "insert into settings(key, value_json) values ('current_batch_id', ?) "
            "on conflict(key) do update set value_json = excluded.value_json",
            (str(batch_id),),
        )
        ledger_counts: dict[str, int] = {}
        for ledger_type, (sheet_name, rows) in parsed.items():
            ledger_counts[ledger_type] = len(rows)
            for row_number, row in rows:
                row_json = json.dumps(row, ensure_ascii=False, default=str)
                raw_row_id = conn.execute(
                    "insert into raw_rows(batch_id, ledger_type, sheet_name, row_number, row_json) values (?, ?, ?, ?, ?)",
                    (batch_id, ledger_type, sheet_name, row_number, row_json),
                ).lastrowid
                conn.execute(
                    """
                    insert into ledger_rows(
                        batch_id, ledger_type, city, district, telecom_site_code, telecom_site_name,
                        tower_site_code, tower_site_name, raw_row_id, row_json, sheet_name, row_number
                    ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        ledger_type,
                        _clean(row.get("地市")),
                        _clean(row.get("区县")),
                        _clean(row.get("电信站址编码")),
                        _clean(row.get("电信站址名称")),
                        _clean(row.get("铁塔站址编码")),
                        _clean(row.get("铁塔站址名称")),
                        raw_row_id,
                        "{}",
                        sheet_name,
                        row_number,
                    ),
                )
        total_records = sum(ledger_counts.values())
        elapsed = perf_counter() - started_at
        conn.execute(
            "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
            (batch_id, operation, f"{message}；记录 {total_records} 条；耗时 {elapsed:.2f} 秒"),
        )
    record_recent_file(config, workbook_path, "import", True, ledger_counts, 0)
    return ImportResult(batch_id=batch_id, ledger_counts=ledger_counts)


def _clear_batch_data(conn, batch_id: int) -> None:
    audit_run_ids = [row["id"] for row in conn.execute("select id from audit_runs where batch_id = ?", (batch_id,))]
    conn.execute("delete from issues where batch_id = ?", (batch_id,))
    if audit_run_ids:
        conn.executemany("delete from audit_results where audit_run_id = ?", [(audit_run_id,) for audit_run_id in audit_run_ids])
    conn.execute("delete from audit_runs where batch_id = ?", (batch_id,))
    conn.execute("delete from ledger_rows where batch_id = ?", (batch_id,))
    conn.execute("delete from raw_rows where batch_id = ?", (batch_id,))


def _headers(ws: Worksheet, header_rows: int, ledger_type: LedgerType) -> list[str]:
    if header_rows == 1:
        raw_headers = tuple(cell.value for cell in ws[1])
        return [header for cell in ws[1] if (header := canonical_header(cell.value, ledger_type, raw_headers))]
    first = _parent_headers(ws, ledger_type)
    raw_second_headers = tuple(cell.value for cell in ws[2])
    second = [canonical_header(cell.value, ledger_type, raw_second_headers) for cell in ws[2]]
    headers: list[str] = []
    for parent, child in zip(first, second, strict=False):
        if parent == "发电时间" and child and child != "发电时长":
            headers.append(f"{parent} - {child}")
        elif child:
            headers.append(child)
        elif parent:
            headers.append(parent)
    return headers


def _data_rows(ws: Worksheet, headers: list[str], first_data_row: int) -> list[tuple[int, dict[str, Any]]]:
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
        row = {header: value for header, value in zip(headers, values, strict=False)}
        if any(value not in (None, "") for value in row.values()):
            rows.append((row_number, row))
    return rows


def _parent_headers(ws: Worksheet, ledger_type: LedgerType) -> list[str | None]:
    headers: list[str | None] = []
    raw_headers = tuple(cell.value for cell in ws[1])
    for cell in ws[1]:
        value = canonical_header(cell.value, ledger_type, raw_headers)
        if value:
            headers.append(value)
            continue
        headers.append(_merged_parent_value(ws, cell.column))
    return headers


def _merged_parent_value(ws: Worksheet, column: int) -> str | None:
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row == 1 and cell_range.max_row == 1 and cell_range.min_col <= column <= cell_range.max_col:
            return canonical_header(ws.cell(row=1, column=cell_range.min_col).value)
    return None


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_batch_name(value: str) -> str:
    return re.sub(r"^[0-9a-fA-F]{32}-", "", value).strip() or value
