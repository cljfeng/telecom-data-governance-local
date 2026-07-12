import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from governance_app.audit_rules import rule_metadata
from governance_app.config import AppConfig
from governance_app.db import connect
from governance_app.geo import normalize_city
from governance_app.workflow import transition_batch_in_conn

ISSUE_HEADERS = [
    "问题编号",
    "地市",
    "区县",
    "电信站址编码",
    "电信站址名称",
    "台账类型",
    "规则分类",
    "规则编号",
    "规则名称",
    "风险等级",
    "原台账sheet",
    "原始行号",
    "命中字段",
    "原始字段值",
    "问题说明",
    "建议整改方向",
    "整改结果",
    "整改说明",
    "整改后值",
    "附件说明",
]

ANALYSIS_CORRECTION_HEADERS = [
    "问题编号",
    "整改结果",
    "整改说明",
    "整改后值",
    "机会编号",
    "核实可追回金额",
    "实际落实金额",
]


def export_issue_packages(config: AppConfig, batch_id: int, mode: str = "city") -> list[Path]:
    if mode not in {"city", "province"}:
        raise ValueError("invalid export mode")
    if mode == "province":
        return _export_province_issue_package(config, batch_id)
    return export_city_issue_packages(config, batch_id)


def export_city_issue_packages(config: AppConfig, batch_id: int) -> list[Path]:
    config.export_dir.mkdir(parents=True, exist_ok=True)
    export_root = config.export_dir.resolve()
    paths: list[Path] = []
    with connect(config) as conn:
        batch = conn.execute("select status, is_archived, batch_code from import_batches where id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("batch not found")
        if batch["is_archived"]:
            raise ValueError("batch is archived")
        if batch["status"] != "audited":
            raise ValueError("batch must be audited before export")
        issue_rows = conn.execute(
            _ISSUE_EXPORT_SQL + " order by i.severity, i.issue_code",
            (batch_id,),
        ).fetchall()
        if not issue_rows:
            transition_batch_in_conn(conn, batch_id, "export_empty")
            conn.execute(
                "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
                (batch_id, "export", "当前批次无待导出问题，可直接归档"),
            )
            return []
        grouped: dict[str, list] = {}
        for issue in issue_rows:
            grouped.setdefault(normalize_city(issue["city"]), []).append(issue)
        for city in sorted(grouped):
            issues = grouped[city]
            if not issues:
                continue
            batch_code = batch["batch_code"] or f"批次{batch_id}"
            path = config.export_dir / f"{_safe_filename_part(city)}_整改问题清单_{_safe_filename_part(batch_code)}.xlsx"
            if not path.resolve().is_relative_to(export_root):
                raise ValueError(f"导出路径越界：{path}")
            wb = _issue_workbook()
            ws = wb["整改问题清单"]
            for issue in issues:
                _append_issue_row(ws, issue)
            _finish_issue_sheet(ws)
            wb.save(path)
            conn.executemany(
                """
                update issues
                   set status = 'pending_correction',
                       updated_at = current_timestamp
                 where id = ?
                """,
                [(issue["id"],) for issue in issues],
            )
            conn.executemany(
                """
                insert into issue_events(issue_id, from_status, to_status, source, note)
                values (?, ?, 'pending_correction', 'export', '导出地市整改包')
                """,
                [(issue["id"], issue["status"]) for issue in issues],
            )
            paths.append(path)
        if paths:
            transition_batch_in_conn(conn, batch_id, "export")
            conn.execute(
                "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
                (batch_id, "export", f"导出地市整改包 {len(paths)} 个"),
            )
    return paths


def _export_province_issue_package(config: AppConfig, batch_id: int) -> list[Path]:
    config.export_dir.mkdir(parents=True, exist_ok=True)
    export_root = config.export_dir.resolve()
    with connect(config) as conn:
        batch = conn.execute("select status, is_archived, batch_code from import_batches where id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("batch not found")
        if batch["is_archived"]:
            raise ValueError("batch is archived")
        if batch["status"] != "audited":
            raise ValueError("batch must be audited before export")
        issues = conn.execute(
            _ISSUE_EXPORT_SQL + " order by city, i.severity, i.issue_code",
            (batch_id,),
        ).fetchall()
        if not issues:
            transition_batch_in_conn(conn, batch_id, "export_empty")
            conn.execute(
                "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
                (batch_id, "export", "当前批次无待导出问题，可直接归档"),
            )
            return []
        batch_code = batch["batch_code"] or f"批次{batch_id}"
        path = config.export_dir / f"全省_整改问题清单_{_safe_filename_part(batch_code)}.xlsx"
        if not path.resolve().is_relative_to(export_root):
            raise ValueError(f"导出路径越界：{path}")
        wb = _issue_workbook()
        ws = wb["整改问题清单"]
        for issue in issues:
            _append_issue_row(ws, issue)
        _finish_issue_sheet(ws)
        wb.save(path)
        conn.execute(
            """
            update issues
               set status = 'pending_correction',
                   updated_at = current_timestamp
             where batch_id = ?
            """,
            (batch_id,),
        )
        conn.executemany(
            """
            insert into issue_events(issue_id, from_status, to_status, source, note)
            values (?, ?, 'pending_correction', 'export', '导出全省整改包')
            """,
            [(issue["id"], issue["status"]) for issue in issues],
        )
        transition_batch_in_conn(conn, batch_id, "export")
        conn.execute(
            "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
            (batch_id, "export", "导出全省汇总整改包 1 个"),
        )
    return [path]


def _safe_filename_part(value: str | None) -> str:
    text = str(value or "未填地市").strip()
    text = re.sub(r'[\\/:*?"<>|\s]+', "_", text)
    while ".." in text:
        text = text.replace("..", "_")
    text = text.strip("._")
    return text or "未填地市"


def _issue_workbook() -> Workbook:
    wb = Workbook()
    guide = wb.active
    guide.title = "填写说明"
    guide.append(["整改包填写说明"])
    guide.append(["1. 请逐条核实“整改问题清单”中的问题编号、原始位置、命中字段和问题说明。"])
    guide.append(["2. “整改结果”请选择：已整改、无需整改、情况说明、退回确认。"])
    guide.append(["3. 高风险问题建议在“附件说明”中填写合同、发票、系统截图或现场照片等佐证材料名称。"])
    guide.append(["4. 请勿修改问题编号，系统将按问题编号匹配回传。"])
    ws = wb.create_sheet("整改问题清单")
    ws.append(ISSUE_HEADERS)
    return wb


def _finish_issue_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 18,
        "B": 12,
        "C": 12,
        "D": 18,
        "E": 24,
        "H": 28,
        "I": 24,
        "O": 48,
        "P": 32,
        "Q": 16,
        "R": 32,
        "S": 24,
        "T": 24,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    if ws.max_row < 2:
        return
    validation = DataValidation(type="list", formula1='"已整改,无需整改,情况说明,退回确认"', allow_blank=True)
    validation.error = "请从下拉选项中选择整改结果"
    validation.errorTitle = "整改结果无效"
    validation.prompt = "请选择整改结果"
    validation.promptTitle = "整改结果"
    ws.add_data_validation(validation)
    validation.add(f"Q2:Q{ws.max_row}")


def append_analysis_correction_sheet(wb: Workbook, opportunities: list[dict]) -> None:
    ws = wb.create_sheet("整改问题清单")
    ws.append(ANALYSIS_CORRECTION_HEADERS)
    for opportunity in opportunities:
        if not opportunity["issue_code"]:
            continue
        ws.append(
            [
                excel_safe(opportunity["issue_code"]),
                "",
                excel_safe(opportunity["correction_note"] or opportunity["review_note"]),
                "",
                excel_safe(opportunity["opportunity_code"]),
                opportunity["verified_recoverable_amount"],
                opportunity["realized_saving_amount"],
            ]
        )
    validation = DataValidation(
        type="list",
        formula1='"已整改,无需整改,情况说明,退回确认"',
        allow_blank=True,
    )
    ws.add_data_validation(validation)
    if ws.max_row >= 2:
        validation.add(f"B2:B{ws.max_row}")


def _append_issue_row(ws, issue) -> None:
    metadata = rule_metadata(issue["rule_id"])
    ws.append(
        [
            excel_safe(issue["issue_code"]),
            excel_safe(normalize_city(issue["city"])),
            excel_safe(issue["district"]),
            excel_safe(issue["telecom_site_code"]),
            excel_safe(issue["telecom_site_name"]),
            excel_safe(_ledger_label(issue["ledger_type"])),
            excel_safe(_rule_category_label(metadata.category)),
            excel_safe(issue["rule_id"]),
            excel_safe(metadata.name),
            excel_safe(_severity_label(issue["severity"])),
            excel_safe(issue["sheet_name"]),
            issue["row_number"] or "",
            excel_safe(issue["field_name"]),
            excel_safe(_matched_field_value(issue)),
            excel_safe(_detailed_issue_message(issue)),
            excel_safe(issue["suggestion"]),
            "",
            "",
            "",
            "",
        ]
    )


_ISSUE_EXPORT_SQL = """
            select i.id, i.issue_code, i.audit_result_id, i.batch_id, coalesce(i.city, '未填地市') as city,
                   i.district, i.telecom_site_code, i.telecom_site_name, i.ledger_type, i.rule_id, i.severity,
                   i.status, i.message, i.suggestion, i.correction_value, i.correction_note, i.updated_at,
                   ar.field_name,
                   case
                       when lr.row_json is not null and lr.row_json <> '{}' then lr.row_json
                       else coalesce(rr.row_json, lr.row_json)
                   end as row_json,
                   lr.sheet_name, lr.row_number
              from issues i
              join audit_results ar on ar.id = i.audit_result_id
              left join ledger_rows lr on lr.id = ar.ledger_row_id
              left join raw_rows rr on rr.id = lr.raw_row_id
             where i.batch_id = ?
               and i.status <> 'resolved_by_reaudit'
"""


def _detailed_issue_message(issue) -> str:
    raw = _row_json(issue["row_json"])
    parts = [f"系统在{_ledger_label(issue['ledger_type'])}台账中发现该问题"]
    location = _row_location(issue)
    if location:
        parts.append(location)
    site = _site_sentence(issue)
    if site:
        parts.append(site)
    field_name = issue["field_name"]
    if field_name:
        current_value = raw.get(field_name)
        if current_value not in (None, ""):
            parts.append(f"本次命中的字段是“{field_name}”，原始填报值为“{_format_value(current_value)}”")
        else:
            parts.append(f"本次命中的字段是“{field_name}”，原始填报值为空")
    reason = _format_message(issue["message"])
    if reason:
        parts.append(f"系统判定依据为{reason}")
    parts.append("请在原台账对应行核实后填写整改结果、整改说明和必要附件")
    return "。".join(part.rstrip("。") for part in parts if part) + "。"


def _row_location(issue) -> str:
    sheet_name = issue["sheet_name"]
    row_number = issue["row_number"]
    if sheet_name and row_number:
        return f"原始位置为“{sheet_name}”sheet 第 {row_number} 行"
    if sheet_name:
        return f"原始位置在“{sheet_name}”sheet"
    if row_number:
        return f"原始位置为第 {row_number} 行"
    return ""


def _site_sentence(issue) -> str:
    fragments = []
    if issue["city"]:
        fragments.append(normalize_city(issue["city"]))
    if issue["district"]:
        fragments.append(str(issue["district"]))
    site_name = issue["telecom_site_name"]
    site_code = issue["telecom_site_code"]
    if site_name and site_code:
        fragments.append(f"站址“{site_name}”（编码 {site_code}）")
    elif site_name:
        fragments.append(f"站址“{site_name}”")
    elif site_code:
        fragments.append(f"站址编码 {site_code}")
    return "、".join(fragments) if fragments else ""


def _matched_field_value(issue) -> str:
    field_name = issue["field_name"]
    if not field_name:
        return ""
    raw = _row_json(issue["row_json"])
    value = raw.get(field_name)
    return "" if value in (None, "") else _format_value(value)


def _format_message(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return re.sub(r"(?<![\d.])(-?\d+\.\d{3,})(?![\d.])", lambda match: _format_value(match.group(1)), text)


def _format_value(value: object) -> str:
    if isinstance(value, float | int):
        return f"{float(value):.2f}".rstrip("0").rstrip(".")
    text = str(value).strip()
    try:
        number = float(text.replace(",", "").replace("，", "").rstrip("%"))
    except ValueError:
        return text
    suffix = "%" if text.endswith("%") else ""
    return f"{number:.2f}".rstrip("0").rstrip(".") + suffix


def _row_json(value: str | None) -> dict:
    if not value:
        return {}
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return {}


def _ledger_label(value: str | None) -> str:
    labels = {
        "site": "站址",
        "tower_rent": "铁塔租费",
        "electricity": "电费",
        "generator": "发电费",
        "all": "跨台账",
    }
    return labels.get(str(value or ""), str(value or "未知"))


def _severity_label(value: str | None) -> str:
    labels = {"high": "高", "medium": "中", "low": "低"}
    return labels.get(str(value or ""), str(value or "未知"))


def _rule_category_label(value: str | None) -> str:
    labels = {"data_quality": "基础数据质量", "problem_audit": "问题稽核"}
    return labels.get(str(value or ""), str(value or "未知"))


def excel_safe(value: object) -> object:
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
