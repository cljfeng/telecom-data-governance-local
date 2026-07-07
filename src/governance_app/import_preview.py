import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl import load_workbook

from governance_app.config import AppConfig
from governance_app.importer import _data_rows, _headers
from governance_app.models import LedgerType, ValidationErrorDetail
from governance_app.recent_files import list_recent_files, record_recent_file
from governance_app.templates import EXPECTED_SHEETS, HEADER_ROWS, required_headers_for, workbook_sheet_for


@dataclass(frozen=True)
class ImportPreviewResult:
    ok: bool
    batch_name: str
    ledger_counts: dict[str, int] = field(default_factory=dict)
    errors: list[ValidationErrorDetail] = field(default_factory=list)


def preview_workbook(config: AppConfig, workbook_path: Path) -> ImportPreviewResult:
    wb = load_workbook(workbook_path, data_only=True)
    errors: list[ValidationErrorDetail] = []
    ledger_counts: dict[str, int] = {}

    for canonical_sheet_name, ledger_type in EXPECTED_SHEETS.items():
        sheet_name = workbook_sheet_for(wb.sheetnames, ledger_type)
        if sheet_name is None:
            errors.append(ValidationErrorDetail(0, canonical_sheet_name, "缺少必需 sheet"))
            continue
        ws = wb[sheet_name]
        headers = _headers(ws, HEADER_ROWS[ledger_type], ledger_type)
        missing = [name for name in required_headers_for(ledger_type) if name not in headers]
        for name in missing:
            errors.append(ValidationErrorDetail(1, name, "缺少必需字段"))
        if missing:
            ledger_counts[ledger_type] = 0
            continue
        rows = _data_rows(ws, headers, HEADER_ROWS[ledger_type] + 1)
        ledger_counts[ledger_type] = len(rows)
        errors.extend(_quality_errors(ledger_type, rows))

    result = ImportPreviewResult(
        ok=not errors,
        batch_name=workbook_path.stem,
        ledger_counts=_with_all_ledgers(ledger_counts),
        errors=errors,
    )
    record_recent_file(config, workbook_path, "preview", result.ok, result.ledger_counts, len(result.errors))
    return result


def _with_all_ledgers(counts: dict[LedgerType, int]) -> dict[str, int]:
    return {ledger_type: counts.get(ledger_type, 0) for ledger_type in ("site", "tower_rent", "electricity", "generator")}


NUMERIC_FIELDS = {
    "电费单价",
    "分摊比例(%)",
    "发电时长",
    "需核减时长",
    "核减后时长",
    "非5G金额",
    "5G金额",
    "分摊金额",
    "最终分摊金额",
    "产品服务费合计（元/年）（不含税）",
    "维护费(元/年)",
    "场地费(元/年)",
    "电力引入费(元/年)",
}

DATE_FIELDS = {"发电日期", "停租日期", "协议生效时间"}
PERIOD_FIELDS = {"账单月份"}
YES_NO_FIELDS = {
    "是否报账",
    "是否打包报账站址",
    "是否包干站址",
    "是否有机房",
    "是否拉远",
    "是否有空调",
    "是否有蓄电池",
}
YES_NO_FIELD_VALUES = {
    "是否有机房": {"是", "否", "有", "0", "0.7", "机柜", "迷你机柜", "室分站", "电信机房", "迷你柜", "移动机柜", "机"},
    "是否报账": {"是", "否", "迷你机柜"},
    "是否有蓄电池": {"是", "否", "有", "无", "否（人工派单)", "联通自维站"},
}
NUMERIC_TEXT_VALUES = {
    "/",
    "\\",
    "-",
    "#N/A",
    "无",
    "是",
    "否",
    "包干计价",
    "兰州不区分4/5G",
}
NUMERIC_TEXT_PATTERNS = (
    re.compile(r"^\d+(?:\.\d+)?-\d+(?:\.\d+)?$"),
    re.compile(r".*(?:单价|市价|包月|包年|包干|线损|元/年).*"),
)


def _quality_errors(ledger_type: LedgerType, rows: list[tuple[int, dict[str, Any]]]) -> list[ValidationErrorDetail]:
    errors: list[ValidationErrorDetail] = []
    seen_site_codes: dict[str, int] = {}
    for row_number, row in rows:
        for field_name in NUMERIC_FIELDS:
            if field_name in row and not _is_blank(row[field_name]) and _to_number(row[field_name]) is None and not _is_current_numeric_text(row[field_name]):
                errors.append(ValidationErrorDetail(row_number, field_name, "数字格式异常"))
        for field_name in DATE_FIELDS:
            if field_name in row and not _is_blank(row[field_name]) and not _is_date_like(row[field_name]):
                errors.append(ValidationErrorDetail(row_number, field_name, "日期格式异常"))
        for field_name in PERIOD_FIELDS:
            if field_name in row and not _is_blank(row[field_name]) and not _is_period_like(row[field_name]):
                errors.append(ValidationErrorDetail(row_number, field_name, "账期格式异常"))
        for field_name in YES_NO_FIELDS:
            allowed_values = YES_NO_FIELD_VALUES.get(field_name, {"是", "否"})
            if field_name in row and not _is_blank(row[field_name]) and str(row[field_name]).strip() not in allowed_values:
                errors.append(ValidationErrorDetail(row_number, field_name, "枚举值异常，应为：否、是"))
        percent = _to_number(row.get("分摊比例(%)"))
        if percent is not None and not 0 <= percent <= 100:
            errors.append(ValidationErrorDetail(row_number, "分摊比例(%)", "比例超出 0-100 范围"))
        if ledger_type == "site":
            site_code = str(row.get("电信站址编码") or "").strip()
            if site_code:
                if site_code in seen_site_codes:
                    errors.append(ValidationErrorDetail(row_number, "电信站址编码", "站址编码重复"))
                else:
                    seen_site_codes[site_code] = row_number
    return errors


def _to_number(value: object) -> float | None:
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text.endswith("%"):
            text = text[:-1]
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _is_current_numeric_text(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    return text in NUMERIC_TEXT_VALUES or any(pattern.match(text) for pattern in NUMERIC_TEXT_PATTERNS)


def _is_date_like(value: object) -> bool:
    if isinstance(value, datetime | date):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip().replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _is_period_like(value: object) -> bool:
    if isinstance(value, datetime | date):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip().replace("/", "-").replace(".", "-")
    if re.fullmatch(r"\d{1,2}月", text):
        return True
    for fmt in ("%Y-%m", "%Y%m"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() in {"", "/", "\\", "-", "#N/A"})


def export_preview_errors(config: AppConfig, workbook_path: Path, result: ImportPreviewResult) -> Path:
    error_dir = config.export_dir / "import_errors"
    error_dir.mkdir(parents=True, exist_ok=True)
    path = error_dir / f"{_safe_filename_part(workbook_path.stem)}_导入预检错误.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "导入错误明细"
    ws.append(["来源文件", "行号", "字段名", "错误类型", "处理建议"])
    for error in result.errors:
        ws.append(
            [
                str(workbook_path),
                error.row_number,
                error.field_name,
                error.message,
                _suggestion_for(error),
            ]
        )
    wb.save(path)
    return path


def preview_error_payload(error: ValidationErrorDetail) -> dict[str, Any]:
    severity = classify_preview_error(error)
    return {
        "row_number": error.row_number,
        "field_name": error.field_name,
        "message": error.message,
        "severity": severity,
        "severity_label": "必须修复" if severity == "blocker" else "建议修复",
        "action": _suggestion_for(error),
    }


def preview_error_summary(errors: list[ValidationErrorDetail]) -> dict[str, int]:
    summary = {"blocker": 0, "warning": 0}
    for error in errors:
        summary[classify_preview_error(error)] += 1
    return summary


def classify_preview_error(error: ValidationErrorDetail) -> str:
    if "缺少必需" in error.message or "sheet" in error.message:
        return "blocker"
    return "warning"


def _safe_filename_part(value: str) -> str:
    text = re.sub(r'[\\/:*?"<>|\s]+', "_", value.strip())
    while ".." in text:
        text = text.replace("..", "_")
    return text.strip("._") or "导入预检"


def _suggestion_for(error: ValidationErrorDetail) -> str:
    if "sheet" in error.message:
        return "补充模板中缺失的工作表后重新预检"
    if "字段" in error.message:
        return "按省公司模板补齐字段列名后重新预检"
    return "核对模板内容后重新预检"
