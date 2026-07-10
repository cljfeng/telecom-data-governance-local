from pathlib import Path

from openpyxl import Workbook

from governance_app.analytics import dashboard_summary
from governance_app.audit_rules import rule_metadata
from governance_app.config import AppConfig
from governance_app.db import connect
from governance_app.rule_settings import load_rule_settings
from governance_app.version import version_payload
from governance_app.workflow import city_progress, transition_batch_in_conn

CLOSED_STATUSES = {"closed", "not_required", "resolved_by_reaudit"}


def archive_precheck(config: AppConfig, batch_id: int) -> dict:
    with connect(config) as conn:
        batch = conn.execute("select status, is_archived from import_batches where id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("batch not found")
        status_rows = conn.execute(
            """
            select status, count(*) as count
              from issues
             where batch_id = ?
             group by status
            """,
            (batch_id,),
        ).fetchall()
        status_counts = {row["status"]: row["count"] for row in status_rows}
        open_issue_count = sum(count for status, count in status_counts.items() if status not in CLOSED_STATUSES)
        high_risk_open = conn.execute(
            """
            select count(*) as count
              from issues
             where batch_id = ?
               and severity = 'high'
               and status not in ('closed', 'not_required', 'resolved_by_reaudit')
            """,
            (batch_id,),
        ).fetchone()["count"]
        review_count = conn.execute(
            """
            select count(*) as count
              from issues
             where batch_id = ?
               and status = 'needs_review'
            """,
            (batch_id,),
        ).fetchone()["count"]
    blockers = []
    risk_items = []
    if batch["is_archived"]:
        blockers.append({"type": "archived", "message": "批次已归档，不能重复归档"})
    if batch["status"] != "returning":
        blockers.append({"type": "workflow_status", "message": "批次需要完成导出和回传后再归档"})
    if open_issue_count:
        blockers.append({"type": "open_issues", "message": f"仍有 {open_issue_count} 条问题未闭环"})
    if high_risk_open:
        risk_items.append({"type": "high_risk_open", "message": f"仍有 {high_risk_open} 条高风险问题未闭环"})
    if review_count:
        risk_items.append({"type": "needs_review", "message": f"仍有 {review_count} 条问题等待省公司复核"})
    return {
        "ready": not blockers,
        "batch_status": batch["status"],
        "is_archived": bool(batch["is_archived"]),
        "open_issue_count": open_issue_count,
        "status_counts": status_counts,
        "risk_items": risk_items,
        "blockers": blockers,
    }


def archive_batch(config: AppConfig, batch_id: int) -> Path:
    with connect(config) as conn:
        batch = conn.execute("select status, is_archived from import_batches where id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("batch not found")
        if batch["is_archived"]:
            raise ValueError("batch is archived")
        if batch["status"] != "returning":
            raise ValueError("batch must be ready for archive")

    archive_dir = config.export_dir / f"archive_batch_{batch_id}"
    archive_dir.mkdir(parents=True, exist_ok=True)
    path = archive_dir / f"批次{batch_id}_专项治理归档汇总.xlsx"

    summary = dashboard_summary(config, batch_id)
    progress = city_progress(config, batch_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "归档总览"
    ws.append(["指标", "值"])
    ws.append(["批次号", batch_id])
    ws.append(["台账记录数", sum(int(value or 0) for value in summary["ledger_counts"].values())])
    ws.append(["问题总数", sum(int(row["count"] or 0) for row in summary["issues_by_city"])])
    ws.append(["涉及地市", len(summary["issues_by_city"])])
    ws.append(["闭环率", summary["closure_rate"]])
    ws.append(["未闭环问题数", summary["open_issue_count"]])
    ws.append(["待复核问题数", summary["status_counts"].get("needs_review", 0)])
    ws.append(["仍异常问题数", summary["status_counts"].get("still_invalid", 0)])
    ws.append(["无需整改问题数", summary["status_counts"].get("not_required", 0)])

    ws = wb.create_sheet("规则命中排行")
    ws.append(["规则分类", "规则编号", "规则名称", "命中数"])
    for row in summary["issues_by_rule"]:
        metadata = rule_metadata(row["rule_id"])
        ws.append([_rule_category_label(metadata.category), row["rule_id"], row["rule_name"], row["count"]])

    ws = wb.create_sheet("风险等级分布")
    ws.append(["风险等级", "问题数量"])
    with connect(config) as conn:
        for row in conn.execute(
            """
            select severity, count(*) as count
              from issues
             where batch_id = ?
             group by severity
             order by count desc, severity
            """,
            (batch_id,),
        ):
            ws.append([_severity_label(row["severity"]), row["count"]])

    ws = wb.create_sheet("地市整改进度")
    ws.append(["地市", "问题总数", "待整改", "已回传", "待人工复核", "仍异常", "已关闭", "无需整改", "完成率"])
    for row in progress:
        ws.append(
            [
                row["city"],
                row["total_count"],
                row["pending_count"],
                row["returned_count"],
                row["review_count"],
                row["still_invalid_count"],
                row["closed_count"],
                row["not_required_count"],
                row["completion_rate"],
            ]
        )

    ws = wb.create_sheet("问题清单")
    ws.append(["问题编号", "地市", "区县", "站址编码", "站址名称", "台账类型", "规则分类", "规则编号", "规则名称", "风险", "状态", "问题说明", "整改说明"])
    with connect(config) as conn:
        for issue in conn.execute(
            """
            select issue_code, coalesce(city, '未填地市') as city, district, telecom_site_code,
                   telecom_site_name, ledger_type, rule_id, severity, status, message, correction_note
              from issues
             where batch_id = ?
             order by city, issue_code
            """,
            (batch_id,),
        ):
            metadata = rule_metadata(issue["rule_id"])
            ws.append(
                [
                    issue["issue_code"],
                    issue["city"],
                    issue["district"],
                    issue["telecom_site_code"],
                    issue["telecom_site_name"],
                    _ledger_label(issue["ledger_type"]),
                    _rule_category_label(metadata.category),
                    issue["rule_id"],
                    metadata.name,
                    _severity_label(issue["severity"]),
                    _status_label(issue["status"]),
                    issue["message"],
                    issue["correction_note"],
                ]
            )

        ws = wb.create_sheet("操作日志")
        ws.append(["操作", "说明", "时间"])
        for log in conn.execute(
            """
            select operation, message, created_at
              from operation_logs
             where batch_id = ?
             order by id
            """,
            (batch_id,),
        ):
            ws.append([log["operation"], log["message"], log["created_at"]])

        ws = wb.create_sheet("版本与规则快照")
        ws.append(["项目", "值"])
        for key, value in version_payload().items():
            ws.append([key, value])
        ws.append([])
        ws.append(["规则编号", "规则名称", "规则分类", "风险等级", "是否启用", "阈值配置"])
        settings = load_rule_settings(config)
        for row in conn.execute(
            """
            select rule_id, severity, count(*) as count
              from issues
             where batch_id = ?
             group by rule_id, severity
             order by rule_id
            """,
            (batch_id,),
        ):
            metadata = rule_metadata(row["rule_id"])
            setting = settings.get(row["rule_id"])
            ws.append(
                [
                    row["rule_id"],
                    metadata.name,
                    _rule_category_label(metadata.category),
                    _severity_label(row["severity"]),
                    "是" if setting is None or setting.enabled else "否",
                    "" if setting is None else str(setting.config),
                ]
            )

        ws = wb.create_sheet("未闭环问题")
        ws.append(["问题编号", "地市", "站址编码", "台账类型", "规则分类", "规则名称", "风险", "状态", "问题说明"])
        for issue in conn.execute(
            """
            select issue_code, coalesce(city, '未填地市') as city, telecom_site_code,
                   ledger_type, rule_id, severity, status, message
              from issues
             where batch_id = ?
               and status not in ('closed', 'not_required', 'resolved_by_reaudit')
             order by city, issue_code
            """,
            (batch_id,),
        ):
            metadata = rule_metadata(issue["rule_id"])
            ws.append(
                [
                    issue["issue_code"],
                    issue["city"],
                    issue["telecom_site_code"],
                    _ledger_label(issue["ledger_type"]),
                    _rule_category_label(metadata.category),
                    metadata.name,
                    _severity_label(issue["severity"]),
                    _status_label(issue["status"]),
                    issue["message"],
                ]
            )

        ws = wb.create_sheet("专项复盘")
        ws.append(["复盘项", "值", "建议"])
        ws.append(["闭环率", summary["closure_rate"], "低于100%时不得正式归档"])
        ws.append(["未闭环问题数", summary["open_issue_count"], "优先处理高风险和仍异常问题"])
        ws.append(["待复核问题数", summary["status_counts"].get("needs_review", 0), "复核通过后更新为已关闭或无需整改"])
        ws.append(["仍异常问题数", summary["status_counts"].get("still_invalid", 0), "退回地市继续整改"])
        ws.append([])
        ws.append(["规则编号", "规则名称", "可信度", "命中数", "未闭环", "无需整改", "仍异常", "闭环率"])
        for row in summary.get("rule_effectiveness", []):
            ws.append(
                [
                    row["rule_id"],
                    row["rule_name"],
                    row.get("confidence_label", ""),
                    row["total_count"],
                    row["open_count"],
                    row["not_required_count"],
                    row["still_invalid_count"],
                    row["closure_rate"],
                ]
            )

    wb.save(path)
    with connect(config) as conn:
        transition_batch_in_conn(conn, batch_id, "archive")
        conn.execute(
            "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
            (batch_id, "archive", f"生成归档汇总：{path.name}"),
        )
    return path


def export_notice_report(config: AppConfig, batch_id: int) -> Path:
    config.export_dir.mkdir(parents=True, exist_ok=True)
    summary = dashboard_summary(config, batch_id)
    progress = city_progress(config, batch_id)
    batch_code = _batch_code(config, batch_id)
    path = config.export_dir / f"稽核问题通报_{batch_code}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "通报总览"
    total_issues = sum(int(row["count"] or 0) for row in summary["issues_by_city"])
    ws.append(["指标", "值"])
    ws.append(["批次编码", batch_code])
    ws.append(["问题总数", total_issues])
    ws.append(["涉及地市", len(summary["issues_by_city"])])
    ws.append(["未闭环问题", summary["open_issue_count"]])
    ws.append(["闭环率", summary["closure_rate"]])

    ws = wb.create_sheet("地市问题统计")
    ws.append(["地市", "问题总数", "待整改", "已回传", "待复核", "仍异常", "已关闭", "无需整改", "完成率"])
    for row in progress:
        ws.append(
            [
                row["city"],
                row["total_count"],
                row["pending_count"],
                row["returned_count"],
                row["review_count"],
                row["still_invalid_count"],
                row["closed_count"],
                row["not_required_count"],
                row["completion_rate"],
            ]
        )

    ws = wb.create_sheet("分类统计")
    ws.append(["分类维度", "分类项", "规则编号", "规则名称", "风险等级", "问题数量"])
    for row in summary["issues_by_ledger_type"]:
        ws.append(["台账类型", row.get("ledger_label") or _ledger_label(row["ledger_type"]), "", "", "", row["count"]])
    for row in summary["issues_by_severity"]:
        label = row.get("severity_label") or _severity_label(row["severity"])
        ws.append(["风险等级", label, "", "", label, row["count"]])
    for row in summary["issue_categories"]:
        ws.append([
            "规则分类",
            row.get("category_label") or _rule_category_label(rule_metadata(row["rule_id"]).category),
            row["rule_id"],
            row["rule_name"],
            row.get("severity_label") or _severity_label(row["severity"]),
            row["count"],
        ])

    ws = wb.create_sheet("问题明细")
    ws.append(["问题编号", "地市", "区县", "站址编码", "站址名称", "台账类型", "规则分类", "规则编号", "规则名称", "风险", "状态", "问题说明", "建议整改方向"])
    with connect(config) as conn:
        for issue in conn.execute(
            """
            select issue_code, coalesce(city, '未填地市') as city, district, telecom_site_code,
                   telecom_site_name, ledger_type, rule_id, severity, status, message, suggestion
              from issues
             where batch_id = ?
             order by city, rule_id, issue_code
            """,
            (batch_id,),
        ):
            metadata = rule_metadata(issue["rule_id"])
            ws.append(
                [
                    issue["issue_code"],
                    issue["city"],
                    issue["district"],
                    issue["telecom_site_code"],
                    issue["telecom_site_name"],
                    _ledger_label(issue["ledger_type"]),
                    _rule_category_label(metadata.category),
                    issue["rule_id"],
                    metadata.name,
                    _severity_label(issue["severity"]),
                    _status_label(issue["status"]),
                    issue["message"],
                    issue["suggestion"],
                ]
            )

    wb.save(path)
    with connect(config) as conn:
        conn.execute(
            "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
            (batch_id, "notice_report", f"导出稽核问题通报：{path.name}"),
        )
    return path


def _batch_code(config: AppConfig, batch_id: int) -> str:
    with connect(config) as conn:
        row = conn.execute("select batch_code from import_batches where id = ?", (batch_id,)).fetchone()
        if row is None:
            raise ValueError("batch not found")
        return row["batch_code"] or f"批次{batch_id}"


def _ledger_label(value: str | None) -> str:
    return {
        "site": "站址",
        "tower_rent": "铁塔租费",
        "electricity": "电费",
        "generator": "发电费",
        "all": "跨台账",
    }.get(str(value or ""), str(value or "未知"))


def _severity_label(value: str | None) -> str:
    return {"high": "高", "medium": "中", "low": "低"}.get(str(value or ""), str(value or "未知"))


def _rule_category_label(value: str | None) -> str:
    return {"data_quality": "基础数据质量", "problem_audit": "问题稽核"}.get(str(value or ""), str(value or "未知"))


def _status_label(value: str | None) -> str:
    return {
        "pending_export": "待导出",
        "pending_correction": "待整改",
        "returned": "已回传",
        "still_invalid": "仍异常",
        "needs_review": "待复核",
        "closed": "已关闭",
        "not_required": "无需整改",
        "resolved_by_reaudit": "复核稽核已解除",
    }.get(str(value or ""), str(value or "未知"))
