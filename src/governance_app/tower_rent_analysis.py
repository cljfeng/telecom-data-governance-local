import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from governance_app.analysis_reviews import (
    review_payload_fields,
    review_summary_in_conn,
)
from governance_app.audit_rules import parse_row
from governance_app.config import AppConfig
from governance_app.db import connect
from governance_app.exporter import append_analysis_correction_sheet, excel_safe
from governance_app.geo import normalize_city
from governance_app.rule_fields import (
    AMOUNT_FIELD_KEYWORDS,
    MAINTENANCE_DISCOUNT_FIELDS,
    PERIOD_FIELDS,
    POWER_INTRO_FEE_FIELDS,
    PRODUCT_SERVICE_FEE_FIELDS,
    TOWER_FEE_FIELDS,
)
from governance_app.rule_helpers import _first_value, _number, _period_key

TOWER_RENT_DOMAIN = "tower_rent"

RECOVERABLE_RULE_TYPES = {
    "tower_duplicate_product_service_fee": "重复计费",
    "tower_duplicate_maintenance_fee": "重复计费",
    "tower_duplicate_site_fee": "重复计费",
    "tower_duplicate_power_intro_fee": "重复计费",
    "tower_product_units_zero_fee_nonzero": "产品单元异常",
    "tower_original_owner_power_intro_fee_nonzero": "原产权方费用异常",
    "tower_stopped_site_still_charged": "停租仍计费",
    "tower_charged_after_stop_period": "停租仍计费",
    "fee_paid_without_master_site": "无站址仍付费",
}

DISCOUNT_RULE_TYPES = {
    "tower_maintenance_discount_not_lowest": "共享折扣异常",
}

REVIEW_RULE_TYPES = {
    "tower_product_shared_users_inconsistent": "共享用户数异常",
    "tower_room_shared_users_inconsistent": "共享用户数异常",
    "tower_confirmation_product_changed": "产品属性异常",
    "tower_mount_height_exceeds_tower_height": "基础属性异常",
    "tower_site_height_inconsistent": "基础属性异常",
    "fee_amount_period_spike": "金额环比突变",
}


def run_tower_rent_analysis(config: AppConfig, batch_id: int) -> dict[str, int]:
    with connect(config) as conn:
        batch = conn.execute("select id, status, is_archived from import_batches where id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("批次不存在")
        if batch["is_archived"]:
            raise ValueError("归档批次不允许刷新租费异常分析")
        if batch["status"] not in {"audited", "distributed", "returning"}:
            raise ValueError("请先执行稽核，再生成租费异常分析")
        rent_count = conn.execute(
            "select count(*) as c from ledger_rows where batch_id = ? and ledger_type = 'tower_rent'",
            (batch_id,),
        ).fetchone()["c"]
        if not rent_count:
            raise ValueError("当前批次没有铁塔租费台账，无法生成租费异常分析")

        conn.execute("delete from analysis_opportunities where batch_id = ? and domain = ?", (batch_id, TOWER_RENT_DOMAIN))
        inserted = 0
        for issue in _issue_rows(conn, batch_id):
            row_data = parse_row(issue["row_json"])
            payload = _clue_from_issue(batch_id, issue, row_data)
            if payload is None:
                continue
            conn.execute(
                """
                insert into analysis_opportunities(
                    batch_id, ledger_row_id, domain, opportunity_code, source_issue_code, opportunity_type, severity,
                    city, district, telecom_site_code, telecom_site_name, period, meter_no,
                    current_amount, reference_amount, recoverable_amount, saving_opportunity_amount,
                    confidence, source_rule_ids_json, message, suggestion
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    issue["ledger_row_id"],
                    TOWER_RENT_DOMAIN,
                    payload["opportunity_code"],
                    issue["issue_code"],
                    payload["opportunity_type"],
                    issue["severity"],
                    normalize_city(issue["city"]),
                    issue["district"],
                    issue["telecom_site_code"],
                    issue["telecom_site_name"],
                    payload["period"],
                    "",
                    payload["current_amount"],
                    payload["reference_amount"],
                    payload["recoverable_amount"],
                    payload["saving_opportunity_amount"],
                    payload["confidence"],
                    json.dumps([issue["rule_id"]], ensure_ascii=False),
                    payload["message"],
                    payload["suggestion"],
                ),
            )
            inserted += 1
        conn.execute(
            "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
            (batch_id, "tower_rent_analysis", f"生成租费异常线索 {inserted} 条"),
        )
        return {"clue_count": inserted}


def get_tower_rent_summary(config: AppConfig, batch_id: int) -> dict[str, Any]:
    with connect(config) as conn:
        _require_batch(conn, batch_id)
        ledger = conn.execute(
            """
            select count(*) as row_count,
                   count(distinct telecom_site_code) as site_count
              from ledger_rows
             where batch_id = ? and ledger_type = 'tower_rent'
            """,
            (batch_id,),
        ).fetchone()
        total_amount = 0.0
        for row in conn.execute(
            """
            select case
                       when lr.row_json is not null and lr.row_json <> '{}' then lr.row_json
                       else coalesce(rr.row_json, lr.row_json)
                   end as row_json
              from ledger_rows lr
              left join raw_rows rr on rr.id = lr.raw_row_id
             where lr.batch_id = ? and lr.ledger_type = 'tower_rent'
            """,
            (batch_id,),
        ):
            total_amount += _total_rent_amount(parse_row(row["row_json"]))
        amount_row = conn.execute(
            """
            select count(*) as clue_count,
                   count(distinct telecom_site_code) as abnormal_site_count,
                   coalesce(sum(current_amount), 0) as current_amount,
                   coalesce(sum(recoverable_amount), 0) as recoverable_amount,
                   coalesce(sum(saving_opportunity_amount), 0) as saving_opportunity_amount,
                   sum(case when severity = 'high' then 1 else 0 end) as high_risk_count
              from analysis_opportunities
             where batch_id = ? and domain = ?
            """,
            (batch_id, TOWER_RENT_DOMAIN),
        ).fetchone()
        recoverable = float(amount_row["recoverable_amount"] or 0)
        discount = float(amount_row["saving_opportunity_amount"] or 0)
        current = float(amount_row["current_amount"] or 0)
        summary = {
            "batch_id": batch_id,
            "ledger_row_count": int(ledger["row_count"] or 0),
            "site_count": int(ledger["site_count"] or 0),
            "total_rent_amount": round(total_amount, 2),
            "abnormal_site_count": int(amount_row["abnormal_site_count"] or 0),
            "clue_count": int(amount_row["clue_count"] or 0),
            "recoverable_amount": round(recoverable, 2),
            "discount_realization_amount": round(discount, 2),
            "review_amount": round(max(current - recoverable - discount, 0), 2),
            "high_risk_count": int(amount_row["high_risk_count"] or 0),
            "analysis_generated": bool(
                conn.execute(
                    "select 1 from operation_logs where batch_id = ? and operation = 'tower_rent_analysis' limit 1",
                    (batch_id,),
                ).fetchone()
            ),
            "city_rankings": _city_rankings(conn, batch_id),
            "type_breakdown": _type_breakdown(conn, batch_id),
        }
        return {**summary, **review_summary_in_conn(conn, batch_id, TOWER_RENT_DOMAIN)}


def get_tower_rent_clues(config: AppConfig, batch_id: int, filters: dict[str, str] | None = None) -> list[dict[str, Any]]:
    filters = filters or {}
    where = ["ao.batch_id = ?", "ao.domain = ?"]
    params: list[Any] = [batch_id, TOWER_RENT_DOMAIN]
    for key in ("city", "opportunity_type", "severity", "confidence"):
        value = filters.get(key)
        if value:
            where.append(f"ao.{key} = ?")
            params.append(value)
    status = filters.get("status")
    if status:
        where.append("i.status = ?")
        params.append(status)
    with connect(config) as conn:
        _require_batch(conn, batch_id)
        rows = conn.execute(
            f"""
            select ao.*,
                   i.issue_code, i.status as issue_status, i.correction_value, i.correction_note,
                   r.verified_recoverable_amount, r.realized_saving_amount,
                   r.review_note, r.updated_at as reviewed_at
              from analysis_opportunities ao
              left join issues i on i.issue_code = ao.source_issue_code
              left join analysis_opportunity_reviews r on r.opportunity_code = ao.opportunity_code
             where {" and ".join(where)}
             order by ao.recoverable_amount desc, ao.saving_opportunity_amount desc, ao.current_amount desc, ao.id
            """,
            params,
        ).fetchall()
    return [_clue_payload(row) for row in rows]


def export_tower_rent_clues(config: AppConfig, batch_id: int) -> Path:
    summary = get_tower_rent_summary(config, batch_id)
    if not summary["analysis_generated"]:
        raise ValueError("请先生成租费异常分析，再导出 Excel")
    clues = get_tower_rent_clues(config, batch_id)
    config.export_dir.mkdir(parents=True, exist_ok=True)
    path = config.export_dir / f"批次{batch_id}_租费异常线索清单.xlsx"
    wb = Workbook()
    guide = wb.active
    guide.title = "填写说明"
    guide.append(["租费异常线索清单"])
    guide.append(["预计可追回金额用于相对确定的问题；优惠落实金额用于折扣优惠类线索；待核查金额代表核查范围，不等同于确定损失。"])
    guide.append(["测算金额仅供核查参考，只有核实可追回金额和实际落实金额计入成果。"])
    guide.append(["旧版专题机会请先重新运行专题分析，再进行整改闭环。"])
    ws = wb.create_sheet("异常线索清单")
    ws.append(
        [
            "线索编号",
            "地市",
            "区县",
            "站址编码",
            "站址名称",
            "账期",
            "异常类型",
            "风险等级",
            "当前金额",
            "参考金额",
            "预计可追回金额",
            "优惠落实金额",
            "待核查金额",
            "置信度",
            "来源规则",
            "问题说明",
            "建议动作",
        ]
    )
    for item in clues:
        ws.append(
            [
                excel_safe(item["opportunity_code"]),
                excel_safe(item["city"]),
                excel_safe(item["district"]),
                excel_safe(item["telecom_site_code"]),
                excel_safe(item["telecom_site_name"]),
                excel_safe(item["period"]),
                excel_safe(item["opportunity_type"]),
                excel_safe(item["severity"]),
                item["current_amount"],
                item["reference_amount"],
                item["recoverable_amount"],
                item["discount_realization_amount"],
                item["review_amount"],
                excel_safe(item["confidence"]),
                excel_safe(",".join(item["source_rule_ids"])),
                excel_safe(item["message"]),
                excel_safe(item["suggestion"]),
            ]
        )
    city = wb.create_sheet("地市汇总")
    city.append(["地市", "线索数量", "预计可追回金额", "优惠落实金额", "待核查金额"])
    for item in summary["city_rankings"]:
        city.append([excel_safe(item["city"]), item["clue_count"], item["recoverable_amount"], item["discount_realization_amount"], item["review_amount"]])
    type_ws = wb.create_sheet("异常分类汇总")
    type_ws.append(["异常类型", "线索数量", "预计可追回金额", "优惠落实金额", "待核查金额"])
    for item in summary["type_breakdown"]:
        type_ws.append([excel_safe(item["opportunity_type"]), item["clue_count"], item["recoverable_amount"], item["discount_realization_amount"], item["review_amount"]])
    append_analysis_correction_sheet(wb, clues)
    wb.save(path)
    return path


def _issue_rows(conn, batch_id: int):
    return conn.execute(
        """
        select i.id, i.issue_code, i.rule_id, i.severity, i.city, i.district, i.telecom_site_code, i.telecom_site_name,
               i.message, i.suggestion, ar.ledger_row_id,
               case
                   when lr.row_json is not null and lr.row_json <> '{}' then lr.row_json
                   else coalesce(rr.row_json, lr.row_json)
               end as row_json
          from issues i
          join audit_results ar on ar.id = i.audit_result_id
          join ledger_rows lr on lr.id = ar.ledger_row_id
          left join raw_rows rr on rr.id = lr.raw_row_id
         where i.batch_id = ? and i.ledger_type = 'tower_rent'
           and i.status <> 'resolved_by_reaudit'
         order by i.id
        """,
        (batch_id,),
    ).fetchall()


def _clue_from_issue(batch_id: int, issue, row: dict[str, Any]) -> dict[str, Any] | None:
    rule_id = issue["rule_id"]
    opportunity_type = RECOVERABLE_RULE_TYPES.get(rule_id) or DISCOUNT_RULE_TYPES.get(rule_id) or REVIEW_RULE_TYPES.get(rule_id)
    if opportunity_type is None:
        return None
    current_amount = _amount_for_rule(rule_id, row)
    recoverable = current_amount if rule_id in RECOVERABLE_RULE_TYPES else 0.0
    discount = _discount_amount(rule_id, current_amount, row)
    confidence = "high" if rule_id in RECOVERABLE_RULE_TYPES else "medium"
    if rule_id in REVIEW_RULE_TYPES:
        confidence = "medium" if current_amount else "low"
    return {
        "opportunity_code": _clue_code(batch_id, issue["ledger_row_id"], opportunity_type, rule_id),
        "opportunity_type": opportunity_type,
        "period": _period(row),
        "current_amount": current_amount,
        "reference_amount": round(max(current_amount - recoverable - discount, 0), 2),
        "recoverable_amount": recoverable,
        "saving_opportunity_amount": discount,
        "confidence": confidence,
        "message": issue["message"],
        "suggestion": _suggestion(rule_id, issue["suggestion"]),
    }


def _amount_for_rule(rule_id: str, row: dict[str, Any]) -> float:
    field_map = {
        "tower_duplicate_product_service_fee": PRODUCT_SERVICE_FEE_FIELDS,
        "tower_duplicate_maintenance_fee": ("维护费(元/年)", "维护费"),
        "tower_duplicate_site_fee": ("场地费(元/年)", "场地费"),
        "tower_duplicate_power_intro_fee": POWER_INTRO_FEE_FIELDS,
        "tower_original_owner_power_intro_fee_nonzero": POWER_INTRO_FEE_FIELDS,
        "tower_product_units_zero_fee_nonzero": PRODUCT_SERVICE_FEE_FIELDS,
    }
    fields = field_map.get(rule_id)
    if fields:
        return _money(_first_value(row, fields))
    return _total_rent_amount(row)


def _discount_amount(rule_id: str, current_amount: float, row: dict[str, Any]) -> float:
    if rule_id != "tower_maintenance_discount_not_lowest":
        return 0.0
    maintenance = _money(_first_value(row, ("维护费(元/年)", "维护费")))
    discount = _number(_first_value(row, MAINTENANCE_DISCOUNT_FIELDS))
    if maintenance <= 0 or discount is None or float(discount) <= 0:
        return 0.0
    reference_discount = 0.7
    return round(max(maintenance - (maintenance / float(discount)) * reference_discount, 0), 2)


def _total_rent_amount(row: dict[str, Any]) -> float:
    total = 0.0
    for field in TOWER_FEE_FIELDS:
        total += _money(row.get(field))
    if total:
        return round(total, 2)
    for key, value in row.items():
        if any(keyword in str(key) for keyword in AMOUNT_FIELD_KEYWORDS):
            total += _money(value)
    return round(total, 2)


def _require_batch(conn, batch_id: int) -> None:
    if conn.execute("select 1 from import_batches where id = ?", (batch_id,)).fetchone() is None:
        raise ValueError("批次不存在")


def _clue_payload(row) -> dict[str, Any]:
    payload = dict(row)
    payload["source_rule_ids"] = json.loads(payload.pop("source_rule_ids_json") or "[]")
    for field in ("current_amount", "reference_amount", "recoverable_amount", "saving_opportunity_amount"):
        payload[field] = round(float(payload[field] or 0), 2)
    payload["discount_realization_amount"] = payload["saving_opportunity_amount"]
    payload["review_amount"] = round(max(payload["current_amount"] - payload["recoverable_amount"] - payload["discount_realization_amount"], 0), 2)
    payload.update(review_payload_fields(row))
    return payload


def _city_rankings(conn, batch_id: int) -> list[dict[str, Any]]:
    return [
        _summary_payload(row, "city")
        for row in conn.execute(
            """
            select coalesce(city, '未填地市') as city,
                   count(*) as clue_count,
                   coalesce(sum(current_amount), 0) as current_amount,
                   coalesce(sum(recoverable_amount), 0) as recoverable_amount,
                   coalesce(sum(saving_opportunity_amount), 0) as saving_opportunity_amount
              from analysis_opportunities
             where batch_id = ? and domain = ?
             group by coalesce(city, '未填地市')
             order by recoverable_amount desc, saving_opportunity_amount desc, current_amount desc
            """,
            (batch_id, TOWER_RENT_DOMAIN),
        )
    ]


def _type_breakdown(conn, batch_id: int) -> list[dict[str, Any]]:
    return [
        _summary_payload(row, "opportunity_type")
        for row in conn.execute(
            """
            select opportunity_type,
                   count(*) as clue_count,
                   coalesce(sum(current_amount), 0) as current_amount,
                   coalesce(sum(recoverable_amount), 0) as recoverable_amount,
                   coalesce(sum(saving_opportunity_amount), 0) as saving_opportunity_amount
              from analysis_opportunities
             where batch_id = ? and domain = ?
             group by opportunity_type
             order by recoverable_amount desc, saving_opportunity_amount desc, current_amount desc
            """,
            (batch_id, TOWER_RENT_DOMAIN),
        )
    ]


def _summary_payload(row, label_field: str) -> dict[str, Any]:
    current = float(row["current_amount"] or 0)
    recoverable = float(row["recoverable_amount"] or 0)
    discount = float(row["saving_opportunity_amount"] or 0)
    label = normalize_city(row[label_field]) if label_field == "city" else row[label_field]
    return {
        label_field: label,
        "clue_count": int(row["clue_count"] or 0),
        "recoverable_amount": round(recoverable, 2),
        "discount_realization_amount": round(discount, 2),
        "review_amount": round(max(current - recoverable - discount, 0), 2),
    }


def _money(value: Any) -> float:
    number = _number(value)
    return round(float(number or 0), 2)


def _period(row: dict[str, Any]) -> str:
    return _period_key(_first_value(row, PERIOD_FIELDS)) or ""


def _clue_code(batch_id: int, ledger_row_id: int, opportunity_type: str, rule_id: str) -> str:
    digest = hashlib.sha1(f"{batch_id}:{ledger_row_id}:{opportunity_type}:{rule_id}".encode("utf-8")).hexdigest()[:10]
    return f"RENT-{batch_id}-{digest}"


def _suggestion(rule_id: str, fallback: str) -> str:
    action = {
        "tower_duplicate_product_service_fee": "核实同账期同站址产品服务费是否重复计费，确认后追回重复费用",
        "tower_duplicate_maintenance_fee": "核实维护费是否重复计费，确认后追回重复费用",
        "tower_duplicate_site_fee": "核实场地费是否重复计费，确认后追回重复费用",
        "tower_duplicate_power_intro_fee": "核实电力引入费是否重复计费，确认后追回重复费用",
        "tower_product_units_zero_fee_nonzero": "核对产品单元数和费用生成口径，产品单元为零时应停止或冲减费用",
        "tower_original_owner_power_intro_fee_nonzero": "核实原产权方站址电力引入费收取依据，确认后冲减或追回",
        "tower_stopped_site_still_charged": "核实停租状态、账期和费用生成口径，确认后停止计费或追回",
        "tower_charged_after_stop_period": "核实停租日期和账期，确认后追回停租后费用",
        "tower_maintenance_discount_not_lowest": "核对共享折扣政策和适用用户数，推动维护费优惠落实",
        "tower_product_shared_users_inconsistent": "核对共享用户数、业务确认单和合同计费参数",
        "tower_room_shared_users_inconsistent": "核对机房共享用户数、业务确认单和合同计费参数",
        "tower_confirmation_product_changed": "核对业务确认单产品变更依据",
        "tower_mount_height_exceeds_tower_height": "核对塔高和挂高基础属性，复核是否影响租费计价",
        "tower_site_height_inconsistent": "统一同站址塔高基础属性，复核是否影响租费计价",
        "fee_paid_without_master_site": "暂停支付并核实站址主数据、费用依据和报账归属",
        "fee_amount_period_spike": "核对账期费用、调账冲销和重复计费原因",
    }.get(rule_id)
    return action or fallback
