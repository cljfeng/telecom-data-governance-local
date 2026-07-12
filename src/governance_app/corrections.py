import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import cast

from openpyxl import load_workbook

from governance_app.analysis_reviews import (
    match_opportunity_in_conn,
    optional_nonnegative_amount,
    sync_existing_review_note_in_conn,
    upsert_review_in_conn,
)
from governance_app.config import AppConfig
from governance_app.db import connect
from governance_app.models import IssueStatus
from governance_app.workflow import (
    transition_batch_in_conn,
    update_issue_status_in_conn,
)


@dataclass(frozen=True)
class CorrectionImportResult:
    matched_count: int
    errors: list[str] = field(default_factory=list)
    review_warnings: list[str] = field(default_factory=list)
    auto_review: dict[str, int] = field(default_factory=dict)


def import_correction_return(config: AppConfig, workbook_path: Path) -> CorrectionImportResult:
    wb = load_workbook(workbook_path, data_only=True)
    if "整改问题清单" not in wb.sheetnames:
        errors = ["缺少 sheet：整改问题清单"]
        _record_return(config, workbook_path, 0, errors, [])
        return CorrectionImportResult(matched_count=0, errors=errors)
    ws = wb["整改问题清单"]
    headers = [cell.value for cell in ws[1]]
    index = {name: pos for pos, name in enumerate(headers)}
    required = ["问题编号", "整改结果", "整改说明", "整改后值"]
    missing = [name for name in required if name not in index]
    if missing:
        errors = [f"缺少回填列：{name}" for name in missing]
        _record_return(config, workbook_path, 0, errors, [])
        return CorrectionImportResult(matched_count=0, errors=errors)
    specialist_headers = ["机会编号", "核实可追回金额", "实际落实金额"]
    present_specialist_headers = [name for name in specialist_headers if name in index]
    if present_specialist_headers and len(present_specialist_headers) != len(
        specialist_headers
    ):
        missing_specialist = [name for name in specialist_headers if name not in index]
        errors = [f"缺少专题回填列：{name}" for name in missing_specialist]
        _record_return(config, workbook_path, 0, errors, [])
        return CorrectionImportResult(matched_count=0, errors=errors)
    is_specialist = len(present_specialist_headers) == len(specialist_headers)

    matched_count = 0
    errors: list[str] = []
    review_warnings: list[str] = []
    auto_review = {"needs_review": 0, "still_invalid": 0, "not_required": 0}
    matched_batch_ids: set[int] = set()
    seen_issue_codes: set[str] = set()
    with connect(config) as conn:
        for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if _is_blank_row(values):
                continue
            issue_code = values[index["问题编号"]]
            if _is_blank(issue_code):
                errors.append(f"第{row_number}行缺少问题编号")
                continue
            result = values[index["整改结果"]]
            note = values[index["整改说明"]]
            corrected = values[index["整改后值"]]
            verified_raw = values[index["核实可追回金额"]] if is_specialist else None
            realized_raw = values[index["实际落实金额"]] if is_specialist else None
            has_specialist_amount = not _is_blank(verified_raw) or not _is_blank(
                realized_raw
            )
            if (
                _is_blank(result)
                and _is_blank(note)
                and _is_blank(corrected)
                and not has_specialist_amount
            ):
                continue
            issue_code_text = str(issue_code)
            if issue_code_text in seen_issue_codes:
                errors.append(f"第{row_number}行问题编号重复：{issue_code_text}")
                continue
            seen_issue_codes.add(issue_code_text)

            verified = None
            realized = None
            row_has_error = False
            if is_specialist:
                try:
                    verified = optional_nonnegative_amount(
                        verified_raw, "核实可追回金额"
                    )
                except ValueError:
                    errors.append(f"第{row_number}行核实可追回金额必须是非负数字")
                    row_has_error = True
                try:
                    realized = optional_nonnegative_amount(realized_raw, "实际落实金额")
                except ValueError:
                    errors.append(f"第{row_number}行实际落实金额必须是非负数字")
                    row_has_error = True
                if has_specialist_amount and _is_blank(result) and _is_blank(note):
                    errors.append(
                        f"第{row_number}行填写专题金额后必须填写整改结果或整改说明"
                    )
                    row_has_error = True

            batch_row = conn.execute(
                """
                select i.id, i.batch_id, i.severity, i.status, b.is_archived
                  from issues i
                  join import_batches b on b.id = i.batch_id
                 where i.issue_code = ?
                """,
                (issue_code_text,),
            ).fetchone()
            if batch_row is not None and batch_row["is_archived"]:
                raise ValueError("batch is archived")
            if batch_row is None and not is_specialist:
                errors.append(f"第{row_number}行问题编号无法匹配：{issue_code}")
                continue

            opportunity = None
            if is_specialist:
                opportunity_code = str(values[index["机会编号"]] or "").strip()
                try:
                    opportunity = match_opportunity_in_conn(
                        conn,
                        opportunity_code,
                        batch_id=(
                            batch_row["batch_id"] if batch_row is not None else None
                        ),
                        expected_issue_code=issue_code_text,
                    )
                except ValueError as exc:
                    errors.append(f"第{row_number}行{exc}")
                    row_has_error = True

            if row_has_error:
                continue
            target_status = _auto_review_status(batch_row, result, note)
            if (
                target_status == "still_invalid"
                and batch_row is not None
                and batch_row["severity"] == "high"
                and _is_blank(note)
            ):
                review_warnings.append(
                    f"第{row_number}行高风险问题缺少整改说明：{issue_code_text}"
                )
            correction_value = None if corrected is None else str(corrected)
            correction_note = None if note is None else str(note or result or "")
            issue_row = update_issue_status_in_conn(
                conn,
                issue_code_text,
                cast(IssueStatus, target_status),
                source="correction_return",
                event_note=f"导入整改回传：{issue_code_text}",
                correction_value=correction_value,
                correction_note=correction_note,
                update_correction_fields=True,
            )
            review_note = str(note or result or "")
            if is_specialist:
                assert opportunity is not None
                upsert_review_in_conn(
                    conn,
                    opportunity,
                    verified,
                    realized,
                    review_note,
                )
            else:
                sync_existing_review_note_in_conn(conn, issue_code_text, review_note)
            matched_batch_ids.add(issue_row["batch_id"])
            auto_review[target_status] = auto_review.get(target_status, 0) + 1
            matched_count += 1
        conn.execute(
            """
            insert into correction_returns(source_file, matched_count, error_count, errors_json, warning_count, warnings_json)
            values (?, ?, ?, ?, ?, ?)
            """,
            (
                str(workbook_path),
                matched_count,
                len(errors),
                json.dumps(errors, ensure_ascii=False),
                len(review_warnings),
                json.dumps(review_warnings, ensure_ascii=False),
            ),
        )
        if matched_batch_ids:
            for batch_id in matched_batch_ids:
                transition_batch_in_conn(conn, batch_id, "correction_return")
                conn.execute(
                    "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
                    (batch_id, "correction_return", f"导入整改回传，匹配 {matched_count} 条"),
                )
    return CorrectionImportResult(
        matched_count=matched_count,
        errors=errors,
        review_warnings=review_warnings,
        auto_review=auto_review,
    )


def _is_blank_row(values: tuple[object, ...]) -> bool:
    return all(_is_blank(value) for value in values)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _auto_review_status(issue_row, result: object, note: object) -> str:
    result_text = str(result or "").strip()
    has_note = not _is_blank(note)
    if "无需整改" in result_text and has_note:
        return "not_required"
    if "退回" in result_text:
        return "still_invalid"
    if issue_row is not None and issue_row["severity"] == "high" and not has_note:
        return "still_invalid"
    return "needs_review"


def _record_return(config: AppConfig, workbook_path: Path, matched_count: int, errors: list[str], warnings: list[str]) -> None:
    with connect(config) as conn:
        conn.execute(
            """
            insert into correction_returns(source_file, matched_count, error_count, errors_json, warning_count, warnings_json)
            values (?, ?, ?, ?, ?, ?)
            """,
            (
                str(workbook_path),
                matched_count,
                len(errors),
                json.dumps(errors, ensure_ascii=False),
                len(warnings),
                json.dumps(warnings, ensure_ascii=False),
            ),
        )
