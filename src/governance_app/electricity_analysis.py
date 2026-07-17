import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from governance_app.analysis_reviews import (
    review_payload_fields,
    review_summary_in_conn,
)
from governance_app.audit_rules import parse_row
from governance_app.config import AppConfig
from governance_app.db import connect
from governance_app.exporter import append_analysis_correction_sheet, excel_safe
from governance_app.geo import normalize_city
from governance_app.rule_fields import (
    ELECTRICITY_AMOUNT_FIELDS,
    PERIOD_FIELDS,
    PRICE_FIELDS,
    USAGE_FIELDS,
)
from governance_app.rule_helpers import _first_value, _number, _period_key, _text

ELECTRICITY_DOMAIN = "electricity"

RECOVERABLE_RULE_TYPES = {
    "electricity_duplicate_payment": "重复报账",
    "electricity_period_overlap": "时段重叠",
    "electricity_zero_usage_positive_fee": "零电量有费用",
    "electricity_amount_calculation_mismatch": "金额计算偏差",
    "electricity_lump_sum_still_reimbursed": "包干重复报账",
}

SAVING_RULE_TYPES = {
    "electricity_price_range": "高电价",
    "electricity_price_commercial_range": "高电价",
    "electricity_price_city_supply_outlier": "高电价",
    "electricity_usage_spike_drop": "异常增长",
    "electricity_transfer_without_contract": "转供电异常",
}


def run_electricity_analysis(config: AppConfig, batch_id: int) -> dict[str, int]:
    with connect(config) as conn:
        batch = conn.execute("select id, status, is_archived from import_batches where id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("批次不存在")
        if batch["is_archived"]:
            raise ValueError("归档批次不允许刷新电费压降分析")
        if batch["status"] not in {"audited", "distributed", "returning"}:
            raise ValueError("请先执行稽核，再生成电费压降分析")
        electricity_count = conn.execute(
            "select count(*) as c from ledger_rows where batch_id = ? and ledger_type = 'electricity'",
            (batch_id,),
        ).fetchone()["c"]
        if not electricity_count:
            raise ValueError("当前批次没有电费台账，无法生成电费压降分析")

        conn.execute("delete from analysis_opportunities where batch_id = ? and domain = ?", (batch_id, ELECTRICITY_DOMAIN))
        inserted = 0
        for issue in _issue_rows(conn, batch_id):
            row_data = parse_row(issue["row_json"])
            payload = _opportunity_from_issue(batch_id, issue, row_data)
            if payload is None:
                continue
            conn.execute(
                """
                insert into analysis_opportunities(
                    batch_id, ledger_row_id, domain, opportunity_code, source_issue_code, opportunity_type, severity,
                    city, district, telecom_site_code, telecom_site_name, period, meter_no,
                    current_amount, reference_amount, recoverable_amount, saving_opportunity_amount,
                    confidence, source_rule_ids_json, message, suggestion
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    issue["ledger_row_id"],
                    ELECTRICITY_DOMAIN,
                    payload["opportunity_code"],
                    issue["issue_code"],
                    payload["opportunity_type"],
                    issue["severity"],
                    normalize_city(issue["city"]),
                    issue["district"],
                    issue["telecom_site_code"],
                    issue["telecom_site_name"],
                    payload["period"],
                    payload["meter_no"],
                    payload["current_amount"],
                    payload["reference_amount"],
                    payload["recoverable_amount"],
                    payload["saving_opportunity_amount"],
                    payload["confidence"],
                    json.dumps([issue["rule_id"]], ensure_ascii=False),
                    payload["message"],
                    payload["suggestion"],
                ),
            )
            inserted += 1
        conn.execute(
            "insert into operation_logs(batch_id, operation, message) values (?, ?, ?)",
            (batch_id, "electricity_analysis", f"生成电费压降机会 {inserted} 条"),
        )
        return {"opportunity_count": inserted}


def get_electricity_summary(config: AppConfig, batch_id: int) -> dict[str, Any]:
    with connect(config) as conn:
        _require_batch(conn, batch_id)
        ledger = conn.execute(
            """
            select count(*) as row_count,
                   count(distinct telecom_site_code) as site_count
              from ledger_rows
             where batch_id = ? and ledger_type = 'electricity'
            """,
            (batch_id,),
        ).fetchone()
        total_amount = 0.0
        for row in conn.execute(
            """
            select case
                       when lr.row_json is not null and lr.row_json <> '{}' then lr.row_json
                       else coalesce(rr.row_json, lr.row_json)
                   end as row_json
              from ledger_rows lr
              left join raw_rows rr on rr.id = lr.raw_row_id
             where lr.batch_id = ? and lr.ledger_type = 'electricity'
            """,
            (batch_id,),
        ):
            total_amount += _first_money(parse_row(row["row_json"]))
        amount_row = conn.execute(
            """
            select count(*) as opportunity_count,
                   count(distinct telecom_site_code) as abnormal_site_count,
                   coalesce(sum(recoverable_amount), 0) as recoverable_amount,
                   coalesce(sum(saving_opportunity_amount), 0) as saving_opportunity_amount,
                   sum(case when severity = 'high' then 1 else 0 end) as high_risk_count
              from analysis_opportunities
             where batch_id = ? and domain = ?
            """,
            (batch_id, ELECTRICITY_DOMAIN),
        ).fetchone()
        generation_exists = bool(
            conn.execute(
                "select 1 from operation_logs where batch_id = ? and operation = 'electricity_analysis' limit 1",
                (batch_id,),
            ).fetchone()
        )
        ledger_row_count = int(ledger["row_count"] or 0)
        summary = {
            "batch_id": batch_id,
            "ledger_row_count": ledger_row_count,
            "site_count": int(ledger["site_count"] or 0),
            "total_electricity_amount": round(total_amount, 2),
            "abnormal_site_count": int(amount_row["abnormal_site_count"] or 0),
            "opportunity_count": int(amount_row["opportunity_count"] or 0),
            "recoverable_amount": round(float(amount_row["recoverable_amount"] or 0), 2),
            "saving_opportunity_amount": round(float(amount_row["saving_opportunity_amount"] or 0), 2),
            "high_risk_count": int(amount_row["high_risk_count"] or 0),
            "analysis_generated": generation_exists and ledger_row_count > 0,
            "analysis_stale": generation_exists and ledger_row_count == 0,
            "city_rankings": _city_rankings(conn, batch_id),
            "type_breakdown": _type_breakdown(conn, batch_id),
        }
        return {**summary, **review_summary_in_conn(conn, batch_id, ELECTRICITY_DOMAIN)}


def get_electricity_opportunities(config: AppConfig, batch_id: int, filters: dict[str, str] | None = None) -> list[dict[str, Any]]:
    filters = filters or {}
    where = ["ao.batch_id = ?", "ao.domain = ?"]
    params: list[Any] = [batch_id, ELECTRICITY_DOMAIN]
    for key in ("city", "opportunity_type", "severity", "confidence"):
        value = filters.get(key)
        if value:
            where.append(f"ao.{key} = ?")
            params.append(value)
    status = filters.get("status")
    if status:
        where.append("i.status = ?")
        params.append(status)
    elif filters.get("queue") == "actionable":
        where.append(
            "i.status in ('pending_export', 'pending_correction', 'returned', 'needs_review', 'still_invalid')"
        )
    if filters.get("review") == "verified":
        where.append("r.verified_recoverable_amount is not null")
    elif filters.get("review") == "realized":
        where.append("r.realized_saving_amount is not null")
    order_by = "ao.recoverable_amount desc, ao.saving_opportunity_amount desc, ao.id"
    if filters.get("queue") == "actionable":
        order_by = """
            case ao.severity when 'high' then 0 when 'medium' then 1 else 2 end,
            case i.status
                when 'needs_review' then 0 when 'returned' then 1
                when 'still_invalid' then 2 when 'pending_correction' then 3
                when 'pending_export' then 4 else 5 end,
            ao.recoverable_amount desc, ao.saving_opportunity_amount desc, ao.id
        """
    with connect(config) as conn:
        _require_batch(conn, batch_id)
        rows = conn.execute(
            f"""
            select ao.*,
                   i.issue_code, i.status as issue_status, i.correction_value, i.correction_note,
                   r.verified_recoverable_amount, r.realized_saving_amount,
                   r.review_note, r.updated_at as reviewed_at
              from analysis_opportunities ao
              left join issues i on i.issue_code = ao.source_issue_code
              left join analysis_opportunity_reviews r on r.opportunity_code = ao.opportunity_code
             where {" and ".join(where)}
             order by {order_by}
            """,
            params,
        ).fetchall()
    return [_opportunity_payload(row) for row in rows]


def export_electricity_opportunities(config: AppConfig, batch_id: int) -> Path:
    summary = get_electricity_summary(config, batch_id)
    if not summary["analysis_generated"]:
        raise ValueError("请先生成电费压降分析，再导出 Excel")
    opportunities = get_electricity_opportunities(config, batch_id)
    config.export_dir.mkdir(parents=True, exist_ok=True)
    path = config.export_dir / f"批次{batch_id}_电费压降机会清单.xlsx"
    wb = Workbook()
    guide = wb.active
    guide.title = "填写说明"
    guide.append(["电费压降机会清单"])
    guide.append(["可追回金额用于相对确定的问题；压降机会金额用于疑似优化空间，不等同于确定损失。"])
    guide.append(["测算金额仅供核查参考，只有核实可追回金额和实际落实金额计入成果。"])
    guide.append(["旧版专题机会请先重新运行专题分析，再进行整改闭环。"])
    ws = wb.create_sheet("机会清单")
    ws.append(
        [
            "机会编号",
            "地市",
            "区县",
            "站址编码",
            "站址名称",
            "电表户号",
            "账期",
            "异常类型",
            "风险等级",
            "当前金额",
            "参考金额",
            "可追回金额",
            "压降机会金额",
            "置信度",
            "来源规则",
            "问题说明",
            "建议动作",
        ]
    )
    for item in opportunities:
        ws.append(
            [
                excel_safe(item["opportunity_code"]),
                excel_safe(item["city"]),
                excel_safe(item["district"]),
                excel_safe(item["telecom_site_code"]),
                excel_safe(item["telecom_site_name"]),
                excel_safe(item["meter_no"]),
                excel_safe(item["period"]),
                excel_safe(item["opportunity_type"]),
                excel_safe(item["severity"]),
                item["current_amount"],
                item["reference_amount"],
                item["recoverable_amount"],
                item["saving_opportunity_amount"],
                excel_safe(item["confidence"]),
                excel_safe(",".join(item["source_rule_ids"])),
                excel_safe(item["message"]),
                excel_safe(item["suggestion"]),
            ]
        )
    city = wb.create_sheet("地市汇总")
    city.append(["地市", "机会数量", "可追回金额", "压降机会金额"])
    for item in summary["city_rankings"]:
        city.append([excel_safe(item["city"]), item["opportunity_count"], item["recoverable_amount"], item["saving_opportunity_amount"]])
    type_ws = wb.create_sheet("异常分类汇总")
    type_ws.append(["异常类型", "机会数量", "可追回金额", "压降机会金额"])
    for item in summary["type_breakdown"]:
        type_ws.append([excel_safe(item["opportunity_type"]), item["opportunity_count"], item["recoverable_amount"], item["saving_opportunity_amount"]])
    append_analysis_correction_sheet(wb, opportunities)
    wb.save(path)
    return path


def _issue_rows(conn, batch_id: int):
    return conn.execute(
        """
        select i.id, i.issue_code, i.rule_id, i.severity, i.city, i.district, i.telecom_site_code, i.telecom_site_name,
               i.message, i.suggestion, ar.ledger_row_id,
               case
                   when lr.row_json is not null and lr.row_json <> '{}' then lr.row_json
                   else coalesce(rr.row_json, lr.row_json)
               end as row_json
          from issues i
          join audit_results ar on ar.id = i.audit_result_id
          join ledger_rows lr on lr.id = ar.ledger_row_id
          left join raw_rows rr on rr.id = lr.raw_row_id
         where i.batch_id = ? and i.ledger_type = 'electricity'
           and i.status <> 'resolved_by_reaudit'
         order by i.id
        """,
        (batch_id,),
    ).fetchall()


def _opportunity_from_issue(batch_id: int, issue, row: dict[str, Any]) -> dict[str, Any] | None:
    rule_id = issue["rule_id"]
    opportunity_type = RECOVERABLE_RULE_TYPES.get(rule_id) or SAVING_RULE_TYPES.get(rule_id)
    if opportunity_type is None:
        return None
    current_amount = _first_money(row)
    recoverable_amount = _recoverable_amount(rule_id, current_amount, row)
    saving_amount = _saving_amount(rule_id, current_amount, row)
    confidence = "high" if rule_id in RECOVERABLE_RULE_TYPES else "medium"
    return {
        "opportunity_code": _opportunity_code(batch_id, issue["ledger_row_id"], opportunity_type, rule_id),
        "opportunity_type": opportunity_type,
        "period": _period(row),
        "meter_no": _meter_no(row),
        "current_amount": current_amount,
        "reference_amount": round(max(current_amount - recoverable_amount - saving_amount, 0), 2),
        "recoverable_amount": recoverable_amount,
        "saving_opportunity_amount": saving_amount,
        "confidence": confidence,
        "message": issue["message"],
        "suggestion": _suggestion(rule_id, issue["suggestion"]),
    }


def _recoverable_amount(rule_id: str, current_amount: float, row: dict[str, Any]) -> float:
    if rule_id in {"electricity_duplicate_payment", "electricity_period_overlap", "electricity_zero_usage_positive_fee", "electricity_lump_sum_still_reimbursed"}:
        return current_amount
    if rule_id == "electricity_amount_calculation_mismatch":
        usage = _usage(row)
        price = _price(row)
        share = _number(row.get("分摊比例(%)"))
        share_factor = (float(share) / 100) if share is not None else 1
        expected = usage * price * share_factor
        return round(max(current_amount - expected, 0), 2)
    return 0.0


def _saving_amount(rule_id: str, current_amount: float, row: dict[str, Any]) -> float:
    if rule_id in {"electricity_price_range", "electricity_price_commercial_range", "electricity_price_city_supply_outlier"}:
        usage = _usage(row)
        price = _price(row)
        reference_price = min(price, 0.9)
        return round(max((price - reference_price) * usage, 0), 2)
    if rule_id in {"electricity_usage_spike_drop", "electricity_transfer_without_contract"}:
        return current_amount
    return 0.0


def _suggestion(rule_id: str, fallback: str) -> str:
    action = {
        "electricity_duplicate_payment": "核实同站址同账期是否重复报账，确认后追回重复支付金额",
        "electricity_period_overlap": "核实抄表起止日期，按重叠周期扣减或追回费用",
        "electricity_zero_usage_positive_fee": "核实零电量费用性质，排除固定费用后追回异常支出",
        "electricity_amount_calculation_mismatch": "复核电量、单价和分摊比例，按差额调整",
        "electricity_lump_sum_still_reimbursed": "核对包干口径，避免包干费用和报账费用重复发生",
        "electricity_price_range": "核实电价依据，推动直供电、合同重谈或转供电降价",
        "electricity_price_commercial_range": "核实电价依据，推动直供电、合同重谈或转供电降价",
        "electricity_price_city_supply_outlier": "对比同区县同供电方式参考价，核实偏高原因",
        "electricity_usage_spike_drop": "核查异常增长原因，排查设备空载、倍率错误和抄表异常",
        "electricity_transfer_without_contract": "补充转供电合同，核实价格依据和加价合理性",
    }.get(rule_id)
    return action or fallback


def _require_batch(conn, batch_id: int) -> None:
    if conn.execute("select 1 from import_batches where id = ?", (batch_id,)).fetchone() is None:
        raise ValueError("批次不存在")


def _opportunity_payload(row) -> dict[str, Any]:
    payload = dict(row)
    payload["source_rule_ids"] = json.loads(payload.pop("source_rule_ids_json") or "[]")
    for field in ("current_amount", "reference_amount", "recoverable_amount", "saving_opportunity_amount"):
        payload[field] = round(float(payload[field] or 0), 2)
    payload.update(review_payload_fields(row))
    return payload


def _city_rankings(conn, batch_id: int) -> list[dict[str, Any]]:
    return [
        {
            "city": normalize_city(row["city"]),
            "opportunity_count": int(row["opportunity_count"] or 0),
            "recoverable_amount": round(float(row["recoverable_amount"] or 0), 2),
            "saving_opportunity_amount": round(float(row["saving_opportunity_amount"] or 0), 2),
        }
        for row in conn.execute(
            """
            select coalesce(city, '未填地市') as city,
                   count(*) as opportunity_count,
                   coalesce(sum(recoverable_amount), 0) as recoverable_amount,
                   coalesce(sum(saving_opportunity_amount), 0) as saving_opportunity_amount
              from analysis_opportunities
             where batch_id = ? and domain = ?
             group by coalesce(city, '未填地市')
             order by recoverable_amount desc, saving_opportunity_amount desc, opportunity_count desc
            """,
            (batch_id, ELECTRICITY_DOMAIN),
        )
    ]


def _type_breakdown(conn, batch_id: int) -> list[dict[str, Any]]:
    return [
        {
            "opportunity_type": row["opportunity_type"],
            "opportunity_count": int(row["opportunity_count"] or 0),
            "recoverable_amount": round(float(row["recoverable_amount"] or 0), 2),
            "saving_opportunity_amount": round(float(row["saving_opportunity_amount"] or 0), 2),
        }
        for row in conn.execute(
            """
            select opportunity_type,
                   count(*) as opportunity_count,
                   coalesce(sum(recoverable_amount), 0) as recoverable_amount,
                   coalesce(sum(saving_opportunity_amount), 0) as saving_opportunity_amount
              from analysis_opportunities
             where batch_id = ? and domain = ?
             group by opportunity_type
             order by recoverable_amount desc, saving_opportunity_amount desc, opportunity_count desc
            """,
            (batch_id, ELECTRICITY_DOMAIN),
        )
    ]


def _money(value: Any) -> float:
    number = _number(value)
    return round(float(number or 0), 2)


def _first_money(row: dict[str, Any]) -> float:
    return _money(_first_value(row, ELECTRICITY_AMOUNT_FIELDS))


def _usage(row: dict[str, Any]) -> float:
    return _money(_first_value(row, USAGE_FIELDS))


def _price(row: dict[str, Any]) -> float:
    return _money(_first_value(row, PRICE_FIELDS))


def _period(row: dict[str, Any]) -> str:
    return _period_key(_first_value(row, PERIOD_FIELDS)) or ""


def _meter_no(row: dict[str, Any]) -> str:
    return _text(row.get("电表户号")) or ""


def _opportunity_code(batch_id: int, ledger_row_id: int, opportunity_type: str, rule_id: str) -> str:
    digest = hashlib.sha1(f"{batch_id}:{ledger_row_id}:{opportunity_type}:{rule_id}".encode("utf-8")).hexdigest()[:10]
    return f"OPP-{batch_id}-{digest}"
