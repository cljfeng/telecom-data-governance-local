from openpyxl import Workbook

from governance_app.audit_engine import run_audit
from governance_app.db import connect, initialize_database
from governance_app.import_preview import preview_workbook
from governance_app.importer import import_workbook


def test_large_workbook_preview_import_and_audit_records_observability(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_large_workbook(tmp_path / "large_10000_rows.xlsx", row_count=2500)

    preview = preview_workbook(app_config, workbook_path)
    imported = import_workbook(app_config, workbook_path)
    audit = run_audit(app_config, imported.batch_id)

    assert preview.ok is True
    assert preview.ledger_counts == {"site": 2500, "tower_rent": 2500, "electricity": 2500, "generator": 2500}
    assert imported.ledger_counts == preview.ledger_counts
    assert audit.issue_count == 0
    with connect(app_config) as conn:
        logs = [row["message"] for row in conn.execute("select message from operation_logs order by id")]
        index_names = {row["name"] for row in conn.execute("select name from sqlite_master where type = 'index'")}
    assert any("记录 10000" in message and "耗时" in message for message in logs)
    assert any("生成问题 0 条" in message and "耗时" in message for message in logs)
    assert "idx_ledger_rows_batch_type_city_site" in index_names
    assert "idx_issues_batch_city_status_rule" in index_names


def _save_large_workbook(path, row_count):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("站址台账")
    ws.append(["地市", "区县", "电信站址编码", "电信站址名称", "站址发电责任方"])
    for index in range(row_count):
        ws.append(["杭州", "西湖", f"HZ{index:05d}", f"西湖{index}站", "运营商"])

    ws = wb.create_sheet("铁塔租费台账")
    ws.append(["电信站址编码", "电信站址名称", "地市", "区县", "铁塔站址编码", "铁塔站址名称", "账期"])
    for index in range(row_count):
        ws.append([f"HZ{index:05d}", f"西湖{index}站", "杭州", "西湖", f"TT{index:05d}", f"铁塔西湖{index}站", "2026-04"])

    ws = wb.create_sheet("电费台账")
    ws.append(["地市", "区县", "电信站址编码", "电信站址名称", "电表户号", "报账周期", "电费单价", "供电方式", "分摊比例(%)"])
    for index in range(row_count):
        ws.append(["杭州", "西湖", f"HZ{index:05d}", f"西湖{index}站", f"M{index:05d}", "2026-04", 0.8, "直供电", 100])

    ws = wb.create_sheet("发电费台账")
    ws.append(["序号", "发电日期", "账单月份", "电信站址编码", "电信站址名称", "运维系统工单号", "发电时长"])
    ws.append(["", "", "", "", "", "", ""])
    for index in range(row_count):
        ws.append([index + 1, "2026-04-10", "2026-04", f"HZ{index:05d}", f"西湖{index}站", f"WO{index:05d}", 3])

    wb.save(path)
    return path
