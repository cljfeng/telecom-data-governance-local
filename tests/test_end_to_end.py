import json

from openpyxl import load_workbook

from governance_app.archive import archive_batch
from governance_app.audit_engine import run_audit
from governance_app.corrections import import_correction_return
from governance_app.db import connect, initialize_database
from governance_app.electricity_analysis import (
    export_electricity_opportunities,
    run_electricity_analysis,
)
from governance_app.exporter import export_city_issue_packages
from governance_app.importer import import_workbook
from governance_app.server import create_app


def test_full_local_governance_flow(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    audit = run_audit(app_config, imported.batch_id)
    packages = export_city_issue_packages(app_config, imported.batch_id)

    assert imported.batch_id == 1
    assert audit.audit_run_id == 1
    assert packages

    correction = import_correction_return(app_config, packages[0])
    assert correction.errors == []


def test_specialist_review_closes_through_existing_return_flow_and_archives(
    app_config, sample_workbook
):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        raw = conn.execute(
            """
            select id, row_json
              from raw_rows
             where batch_id = ? and ledger_type = 'electricity'
            """,
            (imported.batch_id,),
        ).fetchone()
        row = json.loads(raw["row_json"])
        row.update(
            {
                "电费单价": 1.2,
                "用电量": 100,
                "电费金额": 300,
                "供电方式": "转供电",
                "转供电合同情况": "无",
            }
        )
        conn.execute(
            "update raw_rows set row_json = ? where id = ?",
            (json.dumps(row, ensure_ascii=False), raw["id"]),
        )

    run_audit(app_config, imported.batch_id)
    run_electricity_analysis(app_config, imported.batch_id)
    specialist_path = export_electricity_opportunities(
        app_config, imported.batch_id
    )
    ordinary_path = export_city_issue_packages(app_config, imported.batch_id)[0]

    ordinary_workbook = load_workbook(ordinary_path)
    ordinary_sheet = ordinary_workbook["整改问题清单"]
    ordinary_headers = {cell.value: cell.column for cell in ordinary_sheet[1]}
    assert "机会编号" not in ordinary_headers
    ordinary_sheet.cell(
        row=2, column=ordinary_headers["整改结果"]
    ).value = "已整改"
    ordinary_sheet.cell(
        row=2, column=ordinary_headers["整改说明"]
    ).value = "普通整改包回传"
    ordinary_workbook.save(ordinary_path)

    ordinary_return = import_correction_return(app_config, ordinary_path)
    assert ordinary_return.errors == []
    with connect(app_config) as conn:
        assert conn.execute(
            "select count(*) from analysis_opportunity_reviews"
        ).fetchone()[0] == 0

    specialist_workbook = load_workbook(specialist_path)
    specialist_sheet = specialist_workbook["整改问题清单"]
    specialist_headers = {cell.value: cell.column for cell in specialist_sheet[1]}
    opportunity_code = specialist_sheet.cell(
        row=2, column=specialist_headers["机会编号"]
    ).value
    specialist_sheet.cell(
        row=2, column=specialist_headers["整改结果"]
    ).value = "已整改"
    specialist_sheet.cell(
        row=2, column=specialist_headers["整改说明"]
    ).value = "已完成专题核查"
    specialist_sheet.cell(
        row=2, column=specialist_headers["核实可追回金额"]
    ).value = 1200.5
    specialist_sheet.cell(
        row=2, column=specialist_headers["实际落实金额"]
    ).value = 800
    specialist_workbook.save(specialist_path)

    specialist_return = import_correction_return(app_config, specialist_path)
    assert specialist_return.errors == []

    response = create_app(app_config).handle_test_request(
        "POST",
        f"/api/batches/{imported.batch_id}/electricity-analysis/review",
        json.dumps(
            {
                "opportunity_code": opportunity_code,
                "status": "closed",
                "verified_recoverable_amount": 1200.5,
                "realized_saving_amount": 800,
                "review_note": "省公司复核通过",
            },
            ensure_ascii=False,
        ),
    )
    assert response[0] == 200
    assert json.loads(response[2])["issue_status"] == "closed"

    with connect(app_config) as conn:
        conn.execute(
            "update issues set status = 'closed' where batch_id = ?",
            (imported.batch_id,),
        )
    run_audit(app_config, imported.batch_id)
    run_electricity_analysis(app_config, imported.batch_id)
    with connect(app_config) as conn:
        refreshed = dict(
            conn.execute(
                """
                select r.opportunity_code, i.status as issue_status,
                       r.verified_recoverable_amount, r.realized_saving_amount
                  from analysis_opportunity_reviews r
                  join issues i on i.issue_code = r.source_issue_code
                 where r.opportunity_code = ?
                """,
                (opportunity_code,),
            ).fetchone()
        )
    precheck = create_app(app_config).handle_test_request(
        "GET", f"/api/archive/precheck?batch_id={imported.batch_id}"
    )
    assert precheck[0] == 200
    assert json.loads(precheck[2])["ready"] is True
    archive_path = archive_batch(app_config, imported.batch_id)

    assert refreshed["opportunity_code"] == opportunity_code
    assert refreshed["issue_status"] == "closed"
    assert refreshed["verified_recoverable_amount"] == 1200.5
    assert refreshed["realized_saving_amount"] == 800.0
    assert load_workbook(archive_path)["专题核查成果"].max_row == 2
