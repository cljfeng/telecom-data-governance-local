import json

import pytest
from openpyxl import load_workbook

from governance_app.analysis_reviews import save_opportunity_review
from governance_app.audit_engine import run_audit
from governance_app.db import connect, initialize_database
from governance_app.tower_rent_analysis import (
    export_tower_rent_clues,
    get_tower_rent_clues,
    get_tower_rent_summary,
    run_tower_rent_analysis,
)


def _create_batch(conn):
    return conn.execute(
        "insert into import_batches(source_file, name, batch_code, status) values (?, ?, ?, ?)",
        ("tower-rent-test.xlsx", "租费测试批次", "BATCH-TOWER", "imported"),
    ).lastrowid


def _insert_row(conn, batch_id, row):
    row_json = json.dumps(row, ensure_ascii=False)
    raw_id = conn.execute(
        "insert into raw_rows(batch_id, ledger_type, sheet_name, row_number, row_json) values (?, 'tower_rent', '铁塔租费台账', ?, ?)",
        (batch_id, row.get("_row_number", 2), row_json),
    ).lastrowid
    return conn.execute(
        """
        insert into ledger_rows(
            batch_id, ledger_type, city, district, telecom_site_code, telecom_site_name,
            tower_site_code, tower_site_name, raw_row_id, row_json, sheet_name, row_number
        ) values (?, 'tower_rent', ?, ?, ?, ?, ?, ?, ?, '{}', '铁塔租费台账', ?)
        """,
        (
            batch_id,
            row.get("地市"),
            row.get("区县"),
            row.get("电信站址编码"),
            row.get("电信站址名称"),
            row.get("铁塔站址编码"),
            row.get("铁塔站址名称"),
            raw_id,
            row.get("_row_number", 2),
        ),
    ).lastrowid


def _audited_tower_batch(app_config):
    initialize_database(app_config)
    with connect(app_config) as conn:
        batch_id = _create_batch(conn)
        rows = [
            {
                "_row_number": 2,
                "地市": "杭州",
                "区县": "西湖",
                "电信站址编码": "T001",
                "电信站址名称": "一站",
                "铁塔站址编码": "TT001",
                "铁塔站址名称": "铁塔一站",
                "账期": "2026-03",
                "订单号": "O001",
                "业务确认单号": "B001",
                "铁塔产品": "普通地面塔A",
                "铁塔共享用户数": 2,
                "机房产品": "自建机房",
                "机房共享用户数": 1,
                "维护费(元/年)": 100,
                "场地费(元/年)": 200,
                "电力引入费(元/年)": 300,
                "产品服务费合计（元/年）（不含税）": 500,
                "铁塔产品单元数": 0,
                "机房产品单元数": 0,
                "配套产品单元数": 0,
                "铁塔共享信息": "共享",
                "维护费共享折扣": 0.9,
                "站址共享信息": "原产权方",
                "挂高": 45,
                "塔高": 40,
            },
            {
                "_row_number": 3,
                "地市": "杭州",
                "区县": "西湖",
                "电信站址编码": "T001",
                "电信站址名称": "一站",
                "铁塔站址编码": "TT001",
                "铁塔站址名称": "铁塔一站",
                "账期": "2026-03",
                "订单号": "O002",
                "业务确认单号": "B001",
                "铁塔产品": "普通地面塔A",
                "铁塔共享用户数": 3,
                "机房产品": "自建机房",
                "机房共享用户数": 2,
                "维护费(元/年)": 50,
                "场地费(元/年)": 100,
                "电力引入费(元/年)": 100,
                "产品服务费合计（元/年）（不含税）": 0,
                "铁塔产品单元数": 1,
                "机房产品单元数": 0,
                "配套产品单元数": 0,
                "铁塔共享信息": "共享",
                "维护费共享折扣": 0.7,
                "站址共享信息": "共享方",
                "挂高": 35,
                "塔高": 35,
            },
        ]
        for row in rows:
            _insert_row(conn, batch_id, row)
    run_audit(app_config, batch_id)
    return batch_id


def test_run_tower_rent_analysis_creates_clues(app_config):
    batch_id = _audited_tower_batch(app_config)

    result = run_tower_rent_analysis(app_config, batch_id)
    clues = get_tower_rent_clues(app_config, batch_id)

    assert result["clue_count"] == len(clues)
    assert result["clue_count"] >= 5
    assert all(item["domain"] == "tower_rent" for item in clues)
    assert any(item["recoverable_amount"] > 0 for item in clues)
    assert any(item["review_amount"] > 0 for item in clues)
    assert clues[0]["issue_code"]
    assert clues[0]["issue_status"] == "pending_export"
    assert clues[0]["verified_recoverable_amount"] is None
    assert get_tower_rent_clues(app_config, batch_id, {"status": "closed"}) == []


def test_run_tower_rent_analysis_refreshes_existing_rows(app_config):
    batch_id = _audited_tower_batch(app_config)

    first = run_tower_rent_analysis(app_config, batch_id)
    second = run_tower_rent_analysis(app_config, batch_id)

    with connect(app_config) as conn:
        count = conn.execute(
            "select count(*) as c from analysis_opportunities where batch_id = ? and domain = 'tower_rent'",
            (batch_id,),
        ).fetchone()["c"]
    assert count == second["clue_count"]
    assert first["clue_count"] == second["clue_count"]


def test_run_tower_rent_analysis_preserves_saved_review(app_config):
    batch_id = _audited_tower_batch(app_config)
    run_tower_rent_analysis(app_config, batch_id)
    before = get_tower_rent_clues(app_config, batch_id)[0]
    save_opportunity_review(
        app_config,
        batch_id,
        "tower-rent-analysis",
        {
            "opportunity_code": before["opportunity_code"],
            "status": "closed",
            "verified_recoverable_amount": 188.5,
            "realized_saving_amount": 32.25,
            "review_note": "已完成租费核查",
        },
    )

    run_tower_rent_analysis(app_config, batch_id)

    after = get_tower_rent_clues(app_config, batch_id, {"status": "closed"})[0]
    assert after["opportunity_code"] == before["opportunity_code"]
    assert after["verified_recoverable_amount"] == 188.5
    assert after["realized_saving_amount"] == 32.25
    assert get_tower_rent_summary(app_config, batch_id)["verified_recoverable_amount"] == 188.5


def test_reaudit_invalidates_stale_tower_rent_analysis(app_config):
    batch_id = _audited_tower_batch(app_config)
    first = run_tower_rent_analysis(app_config, batch_id)
    assert first["clue_count"] > 0

    run_audit(app_config, batch_id)

    assert get_tower_rent_summary(app_config, batch_id)["clue_count"] == 0
    assert get_tower_rent_clues(app_config, batch_id) == []


def test_tower_rent_analysis_ignores_issues_resolved_by_reaudit(app_config):
    batch_id = _audited_tower_batch(app_config)
    with connect(app_config) as conn:
        conn.execute(
            "update issues set status = 'resolved_by_reaudit' where batch_id = ? and ledger_type = 'tower_rent'",
            (batch_id,),
        )

    result = run_tower_rent_analysis(app_config, batch_id)

    assert result["clue_count"] == 0


def test_tower_rent_summary_groups_amounts(app_config):
    batch_id = _audited_tower_batch(app_config)
    run_tower_rent_analysis(app_config, batch_id)

    summary = get_tower_rent_summary(app_config, batch_id)

    assert summary["batch_id"] == batch_id
    assert summary["total_rent_amount"] == 1350
    assert summary["recoverable_amount"] > 0
    assert summary["review_amount"] > 0
    assert summary["analysis_generated"] is True
    assert isinstance(summary["city_rankings"], list)
    assert isinstance(summary["type_breakdown"], list)


def test_tower_rent_export_requires_generated_analysis(app_config):
    batch_id = _audited_tower_batch(app_config)

    assert get_tower_rent_summary(app_config, batch_id)["analysis_generated"] is False
    with pytest.raises(ValueError, match="请先生成租费异常分析"):
        export_tower_rent_clues(app_config, batch_id)


def test_tower_rent_clue_filters(app_config):
    batch_id = _audited_tower_batch(app_config)
    run_tower_rent_analysis(app_config, batch_id)
    all_rows = get_tower_rent_clues(app_config, batch_id)

    filtered = get_tower_rent_clues(
        app_config,
        batch_id,
        filters={"opportunity_type": all_rows[0]["opportunity_type"], "confidence": all_rows[0]["confidence"]},
    )

    assert filtered
    assert all(row["opportunity_type"] == all_rows[0]["opportunity_type"] for row in filtered)
    assert all(row["confidence"] == all_rows[0]["confidence"] for row in filtered)


def test_export_tower_rent_clues_writes_workbook(app_config):
    batch_id = _audited_tower_batch(app_config)
    run_tower_rent_analysis(app_config, batch_id)

    path = export_tower_rent_clues(app_config, batch_id)

    assert path.exists()
    assert "租费异常线索清单" in path.name
    wb = load_workbook(path)
    assert {"填写说明", "异常线索清单", "地市汇总", "异常分类汇总"}.issubset(set(wb.sheetnames))
    headers = [cell.value for cell in wb["异常线索清单"][1]]
    assert "预计可追回金额" in headers
    assert "优惠落实金额" in headers
    assert "待核查金额" in headers
    assert "整改问题清单" in wb.sheetnames
    correction_headers = [cell.value for cell in wb["整改问题清单"][1]]
    assert correction_headers == ["问题编号", "整改结果", "整改说明", "整改后值", "机会编号", "核实可追回金额", "实际落实金额"]


def test_tower_rent_legacy_clue_remains_visible_but_is_not_exported_for_correction(app_config):
    batch_id = _audited_tower_batch(app_config)
    run_tower_rent_analysis(app_config, batch_id)
    with connect(app_config) as conn:
        conn.execute(
            "update analysis_opportunities set source_issue_code = null where batch_id = ? and domain = 'tower_rent'",
            (batch_id,),
        )

    clues = get_tower_rent_clues(app_config, batch_id)
    path = export_tower_rent_clues(app_config, batch_id)
    wb = load_workbook(path)

    assert clues
    assert all(item["issue_code"] is None for item in clues)
    assert wb["整改问题清单"].max_row == 1
    guide_text = "\n".join(str(cell.value or "") for row in wb["填写说明"] for cell in row)
    assert "重新运行专题分析" in guide_text
