from pathlib import Path

import pytest
from openpyxl import Workbook

from governance_app.config import AppConfig


@pytest.fixture
def app_config(tmp_path: Path) -> AppConfig:
    return AppConfig.for_workspace(tmp_path)


def _save_workbook(path: Path, omit_site_code: bool = False) -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    site_headers = ["序号", "地市", "区县", "电信站址编码", "电信站址名称", "经度", "纬度", "站址归属"]
    if omit_site_code:
        site_headers.remove("电信站址编码")
    ws = wb.create_sheet("站址台账")
    ws.append(site_headers)
    ws.append([1, "杭州", "西湖", "HZ001", "西湖一站", 120.1, 30.2, "铁塔"][: len(site_headers)])

    ws = wb.create_sheet("铁塔租费台账")
    ws.append(["序列", "电信站址编码", "电信站址名称", "地市", "区县", "铁塔站址编码", "铁塔站址名称", "产品服务费合计（元/年）（不含税）"])
    ws.append([1, "HZ001", "西湖一站", "杭州", "西湖", "TT001", "铁塔西湖一站", 10000])

    ws = wb.create_sheet("电费台账")
    ws.append(["序号", "地市", "区县", "电信站址编码", "电信站址名称", "电表户号", "报账周期", "电费单价", "供电方式", "分摊比例(%)"])
    ws.append([1, "杭州", "西湖", "HZ001", "西湖一站", "M001", "2026-04", 0.8, "直供电", 100])

    ws = wb.create_sheet("发电费台账")
    ws.append(["序号", "发电日期", "账单月份", "电信站址编码", "电信站址名称", "铁塔站址编码", "铁塔站址名称", "运维系统工单号", "发电时长"])
    ws.append(["", "", "", "", "", "", "", "", ""])
    ws.append([1, "2026-04-10", "2026-04", "HZ001", "西湖一站", "TT001", "铁塔西湖一站", "WO001", 3])

    wb.save(path)
    return path


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    return _save_workbook(tmp_path / "sample_template.xlsx")


@pytest.fixture
def workbook_missing_site_code(tmp_path: Path) -> Path:
    return _save_workbook(tmp_path / "missing_site_code.xlsx", omit_site_code=True)
