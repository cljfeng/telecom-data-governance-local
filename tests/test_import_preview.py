from governance_app.db import connect, initialize_database
from openpyxl import load_workbook

from governance_app.import_preview import export_preview_errors, list_recent_files, preview_workbook


def test_preview_workbook_reports_counts_without_writing_database(app_config, sample_workbook):
    initialize_database(app_config)

    result = preview_workbook(app_config, sample_workbook)

    assert result.ok is True
    assert result.ledger_counts == {"site": 1, "tower_rent": 1, "electricity": 1, "generator": 1}
    assert result.errors == []
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 0
        assert conn.execute("select count(*) as c from ledger_rows").fetchone()["c"] == 0
        assert conn.execute("select count(*) as c from recent_files").fetchone()["c"] == 1


def test_preview_workbook_reports_missing_required_headers(app_config, workbook_missing_site_code):
    initialize_database(app_config)

    result = preview_workbook(app_config, workbook_missing_site_code)

    assert result.ok is False
    assert result.batch_name == "missing_site_code"
    assert result.errors[0].field_name == "电信站址编码"
    assert result.errors[0].message == "缺少必需字段"


def test_export_preview_errors_writes_excel_detail(app_config, workbook_missing_site_code):
    initialize_database(app_config)
    result = preview_workbook(app_config, workbook_missing_site_code)

    path = export_preview_errors(app_config, workbook_missing_site_code, result)

    assert path.exists()
    wb = load_workbook(path)
    ws = wb["导入错误明细"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("来源文件", "行号", "字段名", "错误类型", "处理建议")
    assert rows[1][2] == "电信站址编码"
    assert rows[1][3] == "缺少必需字段"


def test_recent_files_keep_latest_preview_and_import_results(app_config, sample_workbook, workbook_missing_site_code):
    initialize_database(app_config)
    preview_workbook(app_config, workbook_missing_site_code)
    preview_workbook(app_config, sample_workbook)

    recent = list_recent_files(app_config)

    assert [item["path"] for item in recent] == [str(sample_workbook), str(workbook_missing_site_code)]
    assert recent[0]["ok"] is True
    assert recent[0]["ledger_counts"]["site"] == 1
    assert recent[1]["ok"] is False
    assert recent[1]["error_count"] == 1


def test_preview_workbook_reports_row_level_quality_errors(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_quality_error_workbook(tmp_path / "quality_errors.xlsx")

    result = preview_workbook(app_config, workbook_path)

    assert result.ok is False
    errors = {(error.row_number, error.field_name, error.message) for error in result.errors}
    assert (2, "电费单价", "数字格式异常") in errors
    assert (2, "分摊比例(%)", "比例超出 0-100 范围") in errors
    assert (2, "是否报账", "枚举值异常，应为：否、是") in errors
    assert (3, "发电日期", "日期格式异常") in errors
    assert (3, "电信站址编码", "站址编码重复") in errors
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 0


def test_preview_large_workbook_returns_counts_without_writing_database(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_large_workbook(tmp_path / "large_template.xlsx", row_count=250)

    result = preview_workbook(app_config, workbook_path)

    assert result.ok is True
    assert result.ledger_counts == {"site": 250, "tower_rent": 250, "electricity": 250, "generator": 250}
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 0


def _save_quality_error_workbook(path):
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("站址台账")
    ws.append(["地市", "区县", "电信站址编码", "电信站址名称"])
    ws.append(["杭州", "西湖", "HZ001", "西湖一站"])
    ws.append(["杭州", "西湖", "HZ001", "西湖一站重复"])

    ws = wb.create_sheet("铁塔租费台账")
    ws.append(["电信站址编码", "电信站址名称", "地市", "区县", "铁塔站址编码", "铁塔站址名称"])
    ws.append(["HZ001", "西湖一站", "杭州", "西湖", "TT001", "铁塔西湖一站"])

    ws = wb.create_sheet("电费台账")
    ws.append(["地市", "区县", "电信站址编码", "电信站址名称", "电表户号", "报账周期", "电费单价", "分摊比例(%)", "是否报账"])
    ws.append(["杭州", "西湖", "HZ001", "西湖一站", "M001", "2026-04", "abc", 130, "未知"])

    ws = wb.create_sheet("发电费台账")
    ws.append(["序号", "发电日期", "账单月份", "电信站址编码", "电信站址名称", "运维系统工单号", "发电时长"])
    ws.append(["", "", "", "", "", "", ""])
    ws.append([1, "2026-13-40", "2026-04", "HZ001", "西湖一站", "WO001", 3])

    wb.save(path)
    return path


def _save_large_workbook(path, row_count):
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("站址台账")
    ws.append(["地市", "区县", "电信站址编码", "电信站址名称"])
    for index in range(row_count):
        ws.append(["杭州", "西湖", f"HZ{index:05d}", f"西湖{index}站"])

    ws = wb.create_sheet("铁塔租费台账")
    ws.append(["电信站址编码", "电信站址名称", "地市", "区县", "铁塔站址编码", "铁塔站址名称"])
    for index in range(row_count):
        ws.append([f"HZ{index:05d}", f"西湖{index}站", "杭州", "西湖", f"TT{index:05d}", f"铁塔西湖{index}站"])

    ws = wb.create_sheet("电费台账")
    ws.append(["地市", "区县", "电信站址编码", "电信站址名称", "电表户号", "报账周期", "电费单价", "分摊比例(%)"])
    for index in range(row_count):
        ws.append(["杭州", "西湖", f"HZ{index:05d}", f"西湖{index}站", f"M{index:05d}", "2026-04", 0.8, 100])

    ws = wb.create_sheet("发电费台账")
    ws.append(["序号", "发电日期", "账单月份", "电信站址编码", "电信站址名称", "运维系统工单号", "发电时长"])
    ws.append(["", "", "", "", "", "", ""])
    for index in range(row_count):
        ws.append([index + 1, "2026-04-10", "2026-04", f"HZ{index:05d}", f"西湖{index}站", f"WO{index:05d}", 3])

    wb.save(path)
    return path
