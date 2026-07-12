import json

from openpyxl import load_workbook

from governance_app.analysis_reviews import save_opportunity_review
from governance_app.audit_engine import run_audit
from governance_app.db import connect, initialize_database
from governance_app.electricity_analysis import (
    export_electricity_opportunities,
    get_electricity_opportunities,
    get_electricity_summary,
    run_electricity_analysis,
)
from governance_app.importer import import_workbook


def _import_and_audit(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        raw = conn.execute(
            "select id, row_json from raw_rows where batch_id = ? and ledger_type = 'electricity'",
            (imported.batch_id,),
        ).fetchone()
        row = json.loads(raw["row_json"])
        row.update(
            {
                "电费单价": 1.2,
                "用电量": 100,
                "电费金额": 300,
                "供电方式": "转供电",
                "转供电合同情况": "无",
            }
        )
        conn.execute(
            "update raw_rows set row_json = ? where id = ?",
            (json.dumps(row, ensure_ascii=False), raw["id"]),
        )
    run_audit(app_config, imported.batch_id)
    return imported.batch_id


def test_run_electricity_analysis_creates_recoverable_opportunities(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)

    result = run_electricity_analysis(app_config, batch_id)

    opportunities = get_electricity_opportunities(app_config, batch_id)
    assert result["opportunity_count"] == len(opportunities)
    assert result["opportunity_count"] >= 3
    assert any(item["recoverable_amount"] > 0 for item in opportunities)
    assert any(item["saving_opportunity_amount"] > 0 for item in opportunities)
    assert all(item["domain"] == "electricity" for item in opportunities)
    assert all("source_rule_ids" in item for item in opportunities)
    assert opportunities[0]["issue_code"]
    assert opportunities[0]["issue_status"] == "pending_export"
    assert opportunities[0]["verified_recoverable_amount"] is None
    assert get_electricity_opportunities(app_config, batch_id, {"status": "closed"}) == []


def test_run_electricity_analysis_refreshes_existing_rows(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)

    first = run_electricity_analysis(app_config, batch_id)
    second = run_electricity_analysis(app_config, batch_id)

    with connect(app_config) as conn:
        count = conn.execute(
            "select count(*) as c from analysis_opportunities where batch_id = ? and domain = 'electricity'",
            (batch_id,),
        ).fetchone()["c"]
    assert count == second["opportunity_count"]
    assert first["opportunity_count"] == second["opportunity_count"]


def test_run_electricity_analysis_preserves_saved_review(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    run_electricity_analysis(app_config, batch_id)
    before = get_electricity_opportunities(app_config, batch_id)[0]
    save_opportunity_review(
        app_config,
        batch_id,
        "electricity-analysis",
        {
            "opportunity_code": before["opportunity_code"],
            "status": "closed",
            "verified_recoverable_amount": 88.5,
            "realized_saving_amount": 12.25,
            "review_note": "已完成核查",
        },
    )

    run_electricity_analysis(app_config, batch_id)

    after = get_electricity_opportunities(app_config, batch_id, {"status": "closed"})[0]
    assert after["opportunity_code"] == before["opportunity_code"]
    assert after["verified_recoverable_amount"] == 88.5
    assert after["realized_saving_amount"] == 12.25
    assert get_electricity_summary(app_config, batch_id)["verified_recoverable_amount"] == 88.5


def test_reaudit_invalidates_stale_electricity_analysis(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    first = run_electricity_analysis(app_config, batch_id)
    assert first["opportunity_count"] > 0

    run_audit(app_config, batch_id)

    assert get_electricity_summary(app_config, batch_id)["opportunity_count"] == 0
    assert get_electricity_opportunities(app_config, batch_id) == []


def test_electricity_analysis_ignores_issues_resolved_by_reaudit(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update issues set status = 'resolved_by_reaudit' where batch_id = ? and ledger_type = 'electricity'",
            (batch_id,),
        )

    result = run_electricity_analysis(app_config, batch_id)

    assert result["opportunity_count"] == 0


def test_electricity_summary_groups_amounts(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    run_electricity_analysis(app_config, batch_id)

    summary = get_electricity_summary(app_config, batch_id)

    assert summary["batch_id"] == batch_id
    assert summary["total_electricity_amount"] == 300
    assert summary["recoverable_amount"] > 0
    assert summary["saving_opportunity_amount"] > 0
    assert isinstance(summary["city_rankings"], list)
    assert isinstance(summary["type_breakdown"], list)


def test_electricity_opportunity_filters(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    run_electricity_analysis(app_config, batch_id)
    all_rows = get_electricity_opportunities(app_config, batch_id)

    filtered = get_electricity_opportunities(
        app_config,
        batch_id,
        filters={"opportunity_type": all_rows[0]["opportunity_type"], "confidence": all_rows[0]["confidence"]},
    )

    assert filtered
    assert all(row["opportunity_type"] == all_rows[0]["opportunity_type"] for row in filtered)
    assert all(row["confidence"] == all_rows[0]["confidence"] for row in filtered)


def test_export_electricity_opportunities_writes_workbook(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    run_electricity_analysis(app_config, batch_id)

    path = export_electricity_opportunities(app_config, batch_id)

    assert path.exists()
    assert "电费压降机会清单" in path.name
    wb = load_workbook(path)
    assert {"填写说明", "机会清单", "地市汇总", "异常分类汇总"}.issubset(set(wb.sheetnames))
    assert "整改问题清单" in wb.sheetnames
    headers = [cell.value for cell in wb["整改问题清单"][1]]
    assert headers == ["问题编号", "整改结果", "整改说明", "整改后值", "机会编号", "核实可追回金额", "实际落实金额"]


def test_electricity_legacy_opportunity_remains_visible_but_is_not_exported_for_correction(app_config, sample_workbook):
    batch_id = _import_and_audit(app_config, sample_workbook)
    run_electricity_analysis(app_config, batch_id)
    with connect(app_config) as conn:
        conn.execute(
            "update analysis_opportunities set source_issue_code = null where batch_id = ? and domain = 'electricity'",
            (batch_id,),
        )

    opportunities = get_electricity_opportunities(app_config, batch_id)
    path = export_electricity_opportunities(app_config, batch_id)
    wb = load_workbook(path)

    assert opportunities
    assert all(item["issue_code"] is None for item in opportunities)
    assert wb["整改问题清单"].max_row == 1
    guide_text = "\n".join(str(cell.value or "") for row in wb["填写说明"] for cell in row)
    assert "重新运行专题分析" in guide_text
