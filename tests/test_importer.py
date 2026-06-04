import json

from openpyxl import Workbook

from governance_app.audit_engine import run_audit
from governance_app.db import connect, initialize_database
from governance_app.importer import import_workbook


def test_import_workbook_stores_batch_raw_rows_and_ledger_rows(app_config, sample_workbook):
    initialize_database(app_config)

    result = import_workbook(app_config, sample_workbook)

    assert result.batch_id == 1
    assert result.errors == []
    assert result.ledger_counts == {"site": 1, "tower_rent": 1, "electricity": 1, "generator": 1}

    with connect(app_config) as conn:
        raw_count = conn.execute("select count(*) as c from raw_rows").fetchone()["c"]
        ledger_count = conn.execute("select count(*) as c from ledger_rows").fetchone()["c"]
        assert raw_count == 4
        assert ledger_count == 4


def test_import_workbook_reports_missing_required_header(app_config, workbook_missing_site_code):
    initialize_database(app_config)

    result = import_workbook(app_config, workbook_missing_site_code)

    assert result.batch_id is None
    assert result.errors[0].field_name == "电信站址编码"
    assert "缺少必需字段" in result.errors[0].message


def test_import_workbook_accepts_grouped_generator_headers(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_grouped_generator_workbook(tmp_path / "grouped_generator.xlsx")

    result = import_workbook(app_config, workbook_path)

    assert result.errors == []
    with connect(app_config) as conn:
        row_json = conn.execute(
            "select row_json from raw_rows where ledger_type = 'generator'"
        ).fetchone()["row_json"]
    row = json.loads(row_json)
    assert row["发电日期"] == "2026-04-10"
    assert row["发电时间 - 发电开始时间"] == "09:00"
    assert row["发电时间 - 发电结束时间（断电传感器告警消除时间）"] == "12:00"


def test_import_workbook_preserves_excel_row_number_after_blank_rows(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_workbook_with_blank_site_row(tmp_path / "blank_site_row.xlsx")

    result = import_workbook(app_config, workbook_path)

    assert result.errors == []
    with connect(app_config) as conn:
        row_number = conn.execute(
            "select row_number from raw_rows where ledger_type = 'site'"
        ).fetchone()["row_number"]
    assert row_number == 3


def test_import_workbook_accepts_sheet_and_header_aliases(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_alias_workbook(tmp_path / "alias_template.xlsx")

    result = import_workbook(app_config, workbook_path)

    assert result.errors == []
    assert result.ledger_counts == {"site": 1, "tower_rent": 1, "electricity": 1, "generator": 1}
    with connect(app_config) as conn:
        site = conn.execute(
            "select city, telecom_site_code, telecom_site_name from ledger_rows where ledger_type = 'site'"
        ).fetchone()
    assert dict(site) == {
        "city": "杭州",
        "telecom_site_code": "HZ001",
        "telecom_site_name": "西湖一站",
    }


def test_import_workbook_links_ledger_rows_to_raw_rows_without_duplicate_json(app_config, sample_workbook):
    initialize_database(app_config)

    import_workbook(app_config, sample_workbook)

    with connect(app_config) as conn:
        row = conn.execute(
            """
            select lr.raw_row_id, lr.row_json as ledger_json, rr.row_json as raw_json
              from ledger_rows lr
              join raw_rows rr on rr.id = lr.raw_row_id
             where lr.ledger_type = 'electricity'
            """
        ).fetchone()
    assert row["raw_row_id"]
    assert row["ledger_json"] == "{}"
    assert "电费单价" in row["raw_json"]


def test_import_workbook_maps_current_rent_workbook_site_columns(app_config, tmp_path):
    initialize_database(app_config)
    workbook_path = _save_current_rent_format_workbook(tmp_path / "current_rent_format.xlsx")

    result = import_workbook(app_config, workbook_path)

    assert result.errors == []
    with connect(app_config) as conn:
        row = conn.execute(
            """
            select telecom_site_code, telecom_site_name, tower_site_code, tower_site_name
            from ledger_rows
            where ledger_type = 'tower_rent'
            """
        ).fetchone()
    assert dict(row) == {
        "telecom_site_code": "932ADQ.ADQXSF01",
        "telecom_site_name": "安定区新市医院",
        "tower_site_code": "621102908000000451",
        "tower_site_name": "市医院联通共享站",
    }


def test_import_workbook_can_append_to_existing_batch(app_config, sample_workbook):
    initialize_database(app_config)
    first = import_workbook(app_config, sample_workbook)

    second = import_workbook(app_config, sample_workbook, strategy="append", batch_id=first.batch_id)

    assert second.batch_id == first.batch_id
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 1
        assert conn.execute("select count(*) as c from ledger_rows where batch_id = ?", (first.batch_id,)).fetchone()["c"] == 8
        log = conn.execute("select message from operation_logs where operation = 'import_append'").fetchone()
    assert "追加导入" in log["message"]


def test_import_workbook_can_replace_existing_batch_data(app_config, sample_workbook):
    initialize_database(app_config)
    first = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute("update raw_rows set row_json = replace(row_json, '0.8', '9.9') where batch_id = ? and ledger_type = 'electricity'", (first.batch_id,))
    run_audit(app_config, first.batch_id)

    replaced = import_workbook(app_config, sample_workbook, strategy="replace", batch_id=first.batch_id)

    assert replaced.batch_id == first.batch_id
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 1
        assert conn.execute("select count(*) as c from ledger_rows where batch_id = ?", (first.batch_id,)).fetchone()["c"] == 4
        assert conn.execute("select count(*) as c from issues where batch_id = ?", (first.batch_id,)).fetchone()["c"] == 0
        log = conn.execute("select message from operation_logs where operation = 'import_replace'").fetchone()
    assert "覆盖导入" in log["message"]


def _save_grouped_generator_workbook(path):
    wb = _base_workbook()
    ws = wb["发电费台账"]
    ws.delete_rows(1, ws.max_row)
    ws.append(["发电事件", None, None, None, None, None, None, "发电时间", None, "时间时长"])
    ws.append([
        "发电日期",
        "账单月份",
        "电信站址编码",
        "电信站址名称",
        "铁塔站址编码",
        "铁塔站址名称",
        "运维系统工单号",
        "发电开始时间",
        "发电结束时间（断电传感器告警消除时间）",
        "发电时长",
    ])
    ws.merge_cells("A1:G1")
    ws.merge_cells("H1:I1")
    ws.append([
        "2026-04-10",
        "2026-04",
        "HZ001",
        "西湖一站",
        "TT001",
        "铁塔西湖一站",
        "WO001",
        "09:00",
        "12:00",
        3,
    ])
    wb.save(path)
    return path


def _save_workbook_with_blank_site_row(path):
    wb = _base_workbook()
    ws = wb["站址台账"]
    ws.insert_rows(2)
    wb.save(path)
    return path


def _save_alias_workbook(path):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("站址清单")
    ws.append(["序号", "所属地市", "所属区县", "站址编码", "站址名称"])
    ws.append([1, "杭州", "西湖", "HZ001", "西湖一站"])

    ws = wb.create_sheet("铁塔租费清单")
    ws.append(["站址编码", "站址名称", "所属地市", "所属区县", "铁塔编码", "铁塔名称"])
    ws.append(["HZ001", "西湖一站", "杭州", "西湖", "TT001", "铁塔西湖一站"])

    ws = wb.create_sheet("电费清单")
    ws.append(["所属地市", "所属区县", "站址编码", "站址名称", "电表号", "账期"])
    ws.append(["杭州", "西湖", "HZ001", "西湖一站", "M001", "2026-04"])

    ws = wb.create_sheet("发电费清单")
    ws.append(["发电事件", None, None, None, None, "发电时间"])
    ws.append(["发电日期", "账单月份", "站址编码", "站址名称", "工单号", "发电时长"])
    ws.merge_cells("A1:E1")
    ws.append(["2026-04-10", "2026-04", "HZ001", "西湖一站", "WO001", 3])

    wb.save(path)
    return path


def _save_current_rent_format_workbook(path):
    wb = _base_workbook()
    old = wb["铁塔租费台账"]
    wb.remove(old)
    ws = wb.create_sheet("租费台账")
    ws.append([
        "是否存在问题",
        "具体问题",
        "省份",
        "地市",
        "区县",
        "需求单号",
        "运营商",
        "需求承接地市",
        "站址所属地市",
        "需求类型",
        "起租状态",
        "铁塔平台起租状态",
        "业务确认单号",
        "站址编码",
        "站址名称",
        "电信站址编码",
        "电信站址名称",
        "详细地址",
    ])
    ws.append([
        "否",
        "",
        "甘肃省",
        "定西市",
        "安定区",
        "1217052300343238",
        "电信",
        "定西市",
        "定西市",
        "塔类",
        "起租",
        "",
        "CTC-CRM-GSDX-2017-001777",
        "621102908000000451",
        "市医院联通共享站",
        "932ADQ.ADQXSF01",
        "安定区新市医院",
        "甘肃省定西市安定",
    ])
    wb.save(path)
    return path


def _base_workbook():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("站址台账")
    ws.append(["序号", "地市", "区县", "电信站址编码", "电信站址名称", "经度", "纬度", "站址归属"])
    ws.append([1, "杭州", "西湖", "HZ001", "西湖一站", 120.1, 30.2, "铁塔"])

    ws = wb.create_sheet("铁塔租费台账")
    ws.append(["序列", "电信站址编码", "电信站址名称", "地市", "区县", "铁塔站址编码", "铁塔站址名称"])
    ws.append([1, "HZ001", "西湖一站", "杭州", "西湖", "TT001", "铁塔西湖一站"])

    ws = wb.create_sheet("电费台账")
    ws.append(["序号", "地市", "区县", "电信站址编码", "电信站址名称", "电表户号", "报账周期"])
    ws.append([1, "杭州", "西湖", "HZ001", "西湖一站", "M001", "2026-04"])

    ws = wb.create_sheet("发电费台账")
    ws.append(["序号", "发电日期", "账单月份", "电信站址编码", "电信站址名称", "运维系统工单号", "发电时长"])
    ws.append(["", "", "", "", "", "", ""])
    ws.append([1, "2026-04-10", "2026-04", "HZ001", "西湖一站", "WO001", 3])
    return wb
