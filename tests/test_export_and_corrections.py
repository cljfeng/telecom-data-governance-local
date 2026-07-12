import json
import sqlite3

import pytest
from openpyxl import Workbook, load_workbook

from governance_app.analysis_reviews import save_opportunity_review
from governance_app.archive import archive_batch
from governance_app.audit_engine import run_audit
from governance_app.corrections import import_correction_return
from governance_app.db import connect, initialize_database
from governance_app.electricity_analysis import (
    export_electricity_opportunities,
    run_electricity_analysis,
)
from governance_app.exporter import export_city_issue_packages, export_issue_packages
from governance_app.importer import import_workbook


def test_export_city_issue_packages_writes_issue_workbook(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)

    paths = export_city_issue_packages(app_config, imported.batch_id)

    assert paths
    wb = load_workbook(paths[0])
    assert "填写说明" in wb.sheetnames
    assert wb["填写说明"]["A1"].value == "整改包填写说明"
    ws = wb["整改问题清单"]
    assert ws["A1"].value == "问题编号"
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref
    assert ws["F2"].value == "电费"
    assert ws["G1"].value == "规则分类"
    assert ws["G2"].value == "问题稽核"
    assert ws["I1"].value == "规则名称"
    assert ws["I2"].value == "电费高单价"
    assert ws["J2"].value == "高"
    assert ws["K1"].value == "原台账sheet"
    assert ws["L1"].value == "原始行号"
    assert ws["M1"].value == "命中字段"
    assert ws["N1"].value == "原始字段值"
    assert "电费单价" in ws["O2"].value
    assert "当前值" not in ws["O2"].value
    assert "：" not in ws["O2"].value
    assert ws["Q1"].value == "整改结果"
    assert ws.data_validations.count >= 1
    with connect(app_config) as conn:
        export_events = conn.execute(
            "select count(*) as c from issue_events where source = 'export'"
        ).fetchone()["c"]
    assert export_events >= 1


def test_export_issue_packages_can_write_single_province_workbook(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)

    paths = export_issue_packages(app_config, imported.batch_id, mode="province")

    assert len(paths) == 1
    assert "全省" in paths[0].name
    wb = load_workbook(paths[0])
    assert wb.sheetnames == ["填写说明", "整改问题清单"]
    ws = wb["整改问题清单"]
    assert ws["B1"].value == "地市"
    assert ws["B2"].value == "杭州"


def test_export_city_issue_packages_requires_audited_batch(app_config):
    initialize_database(app_config)
    with connect(app_config) as conn:
        batch_id = conn.execute(
            "insert into import_batches(source_file, name, status) values (?, ?, ?)",
            ("manual.xlsx", "未稽核批次", "imported"),
        ).lastrowid

    with pytest.raises(ValueError, match="batch must be audited"):
        export_city_issue_packages(app_config, batch_id)


def test_export_city_issue_packages_marks_no_issue_batch_ready_for_archive(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    run_audit(app_config, imported.batch_id)

    paths = export_city_issue_packages(app_config, imported.batch_id)

    assert paths == []
    with connect(app_config) as conn:
        batch = conn.execute("select status from import_batches where id = ?", (imported.batch_id,)).fetchone()
        log = conn.execute("select message from operation_logs where batch_id = ? order by id desc", (imported.batch_id,)).fetchone()
        assert batch["status"] == "returning"
        assert "无待导出问题" in log["message"]


def test_export_city_issue_packages_sanitizes_city_filename(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    unsafe_city = "杭/州..测试"
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
        conn.execute("update ledger_rows set city = ? where ledger_type = 'electricity'", (unsafe_city,))
    run_audit(app_config, imported.batch_id)

    paths = export_city_issue_packages(app_config, imported.batch_id)

    assert len(paths) == 1
    path = paths[0]
    assert path.resolve().is_relative_to(app_config.export_dir.resolve())
    assert "/" not in path.name
    assert ".." not in path.name
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    assert ws["B2"].value == unsafe_city


def test_export_city_issue_packages_groups_normalized_city_names(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
        conn.execute(
            """
            insert into ledger_rows(
                batch_id, ledger_type, city, district, telecom_site_code, telecom_site_name, row_json
            ) values (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                imported.batch_id,
                "electricity",
                "兰州市",
                "城关",
                "LZ002",
                "兰州二站",
                '{"地市":"兰州市","区县":"城关","电信站址编码":"LZ002","电信站址名称":"兰州二站","电费单价":9.9,"分摊比例(%)":100}',
            ),
        )
        conn.execute("update ledger_rows set city = '兰州' where ledger_type = 'electricity' and telecom_site_code = 'HZ001'")

    run_audit(app_config, imported.batch_id)

    paths = export_city_issue_packages(app_config, imported.batch_id)

    assert len(paths) == 1
    assert paths[0].name.startswith("兰州_整改问题清单_")
    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    exported_cities = {ws.cell(row=row, column=2).value for row in range(2, ws.max_row + 1)}
    assert exported_cities == {"兰州"}


def test_export_city_issue_packages_formats_long_decimal_issue_messages(app_config):
    initialize_database(app_config)
    with connect(app_config) as conn:
        batch_id = conn.execute(
            "insert into import_batches(source_file, name, batch_code, status) values (?, ?, ?, ?)",
            ("manual.xlsx", "长小数批次", "TEST-DECIMAL", "imported"),
        ).lastrowid
        site_json = '{"地市":"兰州","区县":"城关","电信站址编码":"S001","电信站址名称":"已有站"}'
        conn.execute(
            "insert into ledger_rows(batch_id, ledger_type, city, district, telecom_site_code, telecom_site_name, row_json) values (?, ?, ?, ?, ?, ?, ?)",
            (batch_id, "site", "兰州", "城关", "S001", "已有站", site_json),
        )
        fee_json = '{"地市":"兰州市","区县":"城关","电信站址编码":"S003","电信站址名称":"无站付费","产品服务费合计（元/年）（不含税）":100.123456}'
        conn.execute(
            "insert into ledger_rows(batch_id, ledger_type, city, district, telecom_site_code, telecom_site_name, row_json) values (?, ?, ?, ?, ?, ?, ?)",
            (batch_id, "tower_rent", "兰州市", "城关", "S003", "无站付费", fee_json),
        )
    run_audit(app_config, batch_id)

    path = export_city_issue_packages(app_config, batch_id)[0]

    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    messages = [ws.cell(row=row, column=15).value for row in range(2, ws.max_row + 1)]
    fee_message = next(message for message in messages if "无站址仍支付费用" not in message and "正向费用" in message)
    assert "100.12" in fee_message
    assert "100.123" not in fee_message


def test_export_city_issue_packages_escapes_formula_like_values(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
        conn.execute(
            """
            update ledger_rows
               set city = '=HYPERLINK("http://bad")',
                   telecom_site_name = '+SUM(1,1)'
             where ledger_type = 'electricity'
            """
        )
    run_audit(app_config, imported.batch_id)

    path = export_city_issue_packages(app_config, imported.batch_id)[0]

    wb = load_workbook(path, data_only=False)
    ws = wb["整改问题清单"]
    assert ws["B2"].value == "'=HYPERLINK(\"http://bad\")"
    assert ws["B2"].data_type == "s"
    assert ws["E2"].value == "'+SUM(1,1)"
    assert ws["E2"].data_type == "s"


def test_import_correction_return_updates_issue_status(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)

    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    _set_correction_cells(ws, 2, result="已修复", note="已补正")
    wb.save(paths[0])

    result = import_correction_return(app_config, paths[0])

    assert result.matched_count == 1
    with connect(app_config) as conn:
        status = conn.execute("select status from issues limit 1").fetchone()["status"]
        review_count = conn.execute(
            "select count(*) from analysis_opportunity_reviews"
        ).fetchone()[0]
        correction_event = conn.execute(
            "select source from issue_events where source = 'correction_return' limit 1"
        ).fetchone()
        assert status == "needs_review"
        assert review_count == 0
        assert correction_event["source"] == "correction_return"


def test_import_specialist_correction_return_updates_issue_and_review(
    app_config, sample_workbook
):
    path, issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="已完成核查",
        verified=1200.5,
        realized=800,
    )
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 1
    assert result.errors == []
    with connect(app_config) as conn:
        issue = conn.execute(
            "select status, correction_note from issues where issue_code = ?",
            (issue_code,),
        ).fetchone()
        review = conn.execute(
            """
            select verified_recoverable_amount, realized_saving_amount, review_note
              from analysis_opportunity_reviews
            """
        ).fetchone()
    assert tuple(issue) == ("needs_review", "已完成核查")
    assert tuple(review) == (1200.5, 800.0, "已完成核查")


def test_import_specialist_blank_amount_preserves_value_and_zero_overwrites(
    app_config, sample_workbook
):
    path, _, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="首次核查",
        verified=1200.5,
        realized=800,
    )
    wb.save(path)
    assert import_correction_return(app_config, path).errors == []

    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="二次核查",
        verified="",
        realized=0,
    )
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 1
    assert result.errors == []
    with connect(app_config) as conn:
        review = conn.execute(
            """
            select verified_recoverable_amount, realized_saving_amount, review_note
              from analysis_opportunity_reviews
            """
        ).fetchone()
    assert tuple(review) == (1200.5, 0.0, "二次核查")


def test_online_review_preserves_correction_value_imported_from_excel(
    app_config, sample_workbook
):
    path, issue_code, opportunity_code = _specialist_return_workbook(
        app_config, sample_workbook
    )
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    headers = {cell.value: cell.column for cell in ws[1]}
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="Excel 核查完成",
        verified=1200.5,
        realized=800,
    )
    ws.cell(row=2, column=headers["整改后值"]).value = "0.8"
    wb.save(path)
    assert import_correction_return(app_config, path).errors == []

    saved = save_opportunity_review(
        app_config,
        1,
        "electricity-analysis",
        {
            "opportunity_code": opportunity_code,
            "status": "closed",
            "verified_recoverable_amount": 1200.5,
            "realized_saving_amount": 800,
            "review_note": "在线复核通过",
        },
    )

    assert saved["issue_code"] == issue_code
    assert saved["correction_value"] == "0.8"


def test_closed_specialist_export_reimport_without_edits_preserves_status(
    app_config, sample_workbook
):
    _, issue_code, opportunity_code = _specialist_return_workbook(
        app_config, sample_workbook
    )
    save_opportunity_review(
        app_config,
        1,
        "electricity-analysis",
        {
            "opportunity_code": opportunity_code,
            "status": "closed",
            "verified_recoverable_amount": 1200.5,
            "realized_saving_amount": 800,
            "review_note": "省公司复核通过",
        },
    )
    path = export_electricity_opportunities(app_config, 1)
    ws = load_workbook(path)["整改问题清单"]
    headers = {cell.value: cell.column for cell in ws[1]}
    assert ws.cell(row=2, column=headers["整改结果"]).value is None
    assert ws.cell(row=2, column=headers["整改说明"]).value == "省公司复核通过"
    assert ws.cell(row=2, column=headers["核实可追回金额"]).value == 1200.5

    result = import_correction_return(app_config, path)

    assert result.errors == []
    assert result.matched_count == 1
    with connect(app_config) as conn:
        status = conn.execute(
            "select status from issues where issue_code = ?", (issue_code,)
        ).fetchone()[0]
    assert status == "closed"


def test_explicit_specialist_result_can_reopen_closed_issue(
    app_config, sample_workbook
):
    _, issue_code, opportunity_code = _specialist_return_workbook(
        app_config, sample_workbook
    )
    save_opportunity_review(
        app_config,
        1,
        "electricity-analysis",
        {
            "opportunity_code": opportunity_code,
            "status": "closed",
            "review_note": "省公司复核通过",
        },
    )
    path = export_electricity_opportunities(app_config, 1)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers["整改结果"]).value = "退回确认"
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.errors == []
    with connect(app_config) as conn:
        status = conn.execute(
            "select status from issues where issue_code = ?", (issue_code,)
        ).fetchone()[0]
    assert status == "still_invalid"


def test_import_specialist_correction_return_reports_missing_specialist_column(
    app_config, sample_workbook
):
    path, issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=1, column=headers["实际落实金额"]).value = "其他金额"
    _set_specialist_cells(ws, 2, result="已整改", note="已完成核查", verified=1200.5)
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 0
    assert result.errors == ["缺少专题回填列：实际落实金额"]
    with connect(app_config) as conn:
        issue = conn.execute(
            "select status, correction_note from issues where issue_code = ?",
            (issue_code,),
        ).fetchone()
        review_count = conn.execute(
            "select count(*) from analysis_opportunity_reviews"
        ).fetchone()[0]
    assert tuple(issue) == ("pending_correction", None)
    assert review_count == 0


@pytest.mark.parametrize(
    ("header", "value"),
    [("核实可追回金额", -1), ("实际落实金额", "NaN")],
)
def test_import_specialist_correction_return_reports_invalid_amount(
    app_config, sample_workbook, header, value
):
    path, issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(ws, 2, result="已整改", note="已完成核查")
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers[header]).value = value
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 0
    assert result.errors == [f"第2行{header}必须是非负数字"]
    with connect(app_config) as conn:
        status = conn.execute(
            "select status from issues where issue_code = ?", (issue_code,)
        ).fetchone()[0]
        review_count = conn.execute(
            "select count(*) from analysis_opportunity_reviews"
        ).fetchone()[0]
    assert status == "pending_correction"
    assert review_count == 0


def test_import_specialist_correction_return_requires_result_or_note_for_amount(
    app_config, sample_workbook
):
    path, _, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers["核实可追回金额"]).value = 100
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 0
    assert result.errors == ["第2行填写专题金额后必须填写整改结果或整改说明"]


def test_import_specialist_correction_return_reports_opportunity_issue_mismatch(
    app_config, sample_workbook
):
    path, issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="已完成核查",
        verified=100,
    )
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers["问题编号"]).value = "OTHER-ISSUE"
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 0
    assert result.errors == ["第2行专题机会与问题编号不匹配"]
    with connect(app_config) as conn:
        status = conn.execute(
            "select status from issues where issue_code = ?", (issue_code,)
        ).fetchone()[0]
    assert status == "pending_correction"


def test_import_specialist_correction_return_rejects_cross_batch_opportunity_issue(
    app_config, sample_workbook
):
    path, _, opportunity_code = _specialist_return_workbook(
        app_config, sample_workbook
    )
    second = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where batch_id = ? and ledger_type = 'electricity'",
            (second.batch_id,),
        )
    run_audit(app_config, second.batch_id)
    export_city_issue_packages(app_config, second.batch_id)
    with connect(app_config) as conn:
        second_issue_code = conn.execute(
            """
            select issue_code
              from issues
             where batch_id = ? and ledger_type = 'electricity'
             order by id
             limit 1
            """,
            (second.batch_id,),
        ).fetchone()[0]
        conn.execute(
            "update analysis_opportunities set source_issue_code = ? where opportunity_code = ?",
            (second_issue_code, opportunity_code),
        )
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers["问题编号"]).value = second_issue_code
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="跨批次核查",
        verified=100,
    )
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 0
    assert result.errors == ["第2行机会不存在或不属于当前批次专题"]
    with connect(app_config) as conn:
        issue = conn.execute(
            "select status, correction_note from issues where issue_code = ?",
            (second_issue_code,),
        ).fetchone()
        review_count = conn.execute(
            "select count(*) from analysis_opportunity_reviews where opportunity_code = ?",
            (opportunity_code,),
        ).fetchone()[0]
        event_count = conn.execute(
            """
            select count(*)
              from issue_events e
              join issues i on i.id = e.issue_id
             where i.issue_code = ? and e.source = 'correction_return'
            """,
            (second_issue_code,),
        ).fetchone()[0]
    assert tuple(issue) == ("pending_correction", None)
    assert review_count == 0
    assert event_count == 0


def test_import_specialist_correction_return_reports_duplicate_issue_codes(
    app_config, sample_workbook
):
    path, issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(ws, 2, result="已整改", note="已完成核查", verified=100)
    duplicate = [cell.value for cell in ws[2]]
    ws.append(duplicate)
    duplicate_row = ws.max_row
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 1
    assert result.errors == [f"第{duplicate_row}行问题编号重复：{issue_code}"]


def test_import_specialist_correction_return_keeps_valid_rows_when_another_row_is_invalid(
    app_config, sample_workbook
):
    path, first_issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    assert ws.max_row >= 3
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="有效核查",
        verified=1200.5,
        realized=800,
    )
    _set_specialist_cells(ws, 3, result="已整改", note="无效核查", verified=-1)
    second_issue_code = ws.cell(row=3, column=1).value
    assert second_issue_code != first_issue_code
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 1
    assert result.errors == ["第3行核实可追回金额必须是非负数字"]
    with connect(app_config) as conn:
        issues = conn.execute(
            "select issue_code, status from issues where issue_code in (?, ?)",
            (first_issue_code, second_issue_code),
        ).fetchall()
        review_count = conn.execute(
            "select count(*) from analysis_opportunity_reviews"
        ).fetchone()[0]
    assert {row["issue_code"]: row["status"] for row in issues} == {
        first_issue_code: "needs_review",
        second_issue_code: "pending_correction",
    }
    assert review_count == 1


def test_import_normal_correction_return_syncs_existing_review_note(
    app_config, sample_workbook, tmp_path
):
    _, issue_code, opportunity_code = _specialist_return_workbook(
        app_config, sample_workbook
    )
    with connect(app_config) as conn:
        batch_id = conn.execute(
            "select batch_id from issues where issue_code = ?", (issue_code,)
        ).fetchone()[0]
    save_opportunity_review(
        app_config,
        batch_id,
        "electricity-analysis",
        {
            "opportunity_code": opportunity_code,
            "status": "needs_review",
            "verified_recoverable_amount": 10,
            "realized_saving_amount": 5,
            "review_note": "旧说明",
        },
    )
    path = tmp_path / "normal-correction-return.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "整改问题清单"
    ws.append(["问题编号", "整改结果", "整改说明", "整改后值"])
    ws.append([issue_code, "已整改", "普通回传说明", ""])
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 1
    assert result.errors == []
    with connect(app_config) as conn:
        review = conn.execute(
            """
            select verified_recoverable_amount, realized_saving_amount, review_note
              from analysis_opportunity_reviews
             where opportunity_code = ?
            """,
            (opportunity_code,),
        ).fetchone()
    assert tuple(review) == (10.0, 5.0, "普通回传说明")


def test_import_specialist_database_error_rolls_back_whole_import(
    app_config, sample_workbook
):
    path, issue_code, _ = _specialist_return_workbook(app_config, sample_workbook)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    _set_specialist_cells(
        ws,
        2,
        result="已整改",
        note="已完成核查",
        verified=1200.5,
    )
    wb.save(path)
    with connect(app_config) as conn:
        conn.execute(
            """
            create trigger fail_correction_review_insert
            before insert on analysis_opportunity_reviews
            begin
                select raise(abort, 'review insert failed');
            end
            """
        )

    with pytest.raises(sqlite3.IntegrityError, match="review insert failed"):
        import_correction_return(app_config, path)

    with connect(app_config) as conn:
        issue = conn.execute(
            "select status, correction_note from issues where issue_code = ?",
            (issue_code,),
        ).fetchone()
        correction_events = conn.execute(
            "select count(*) from issue_events where source = 'correction_return'"
        ).fetchone()[0]
        review_count = conn.execute(
            "select count(*) from analysis_opportunity_reviews"
        ).fetchone()[0]
        return_count = conn.execute(
            "select count(*) from correction_returns"
        ).fetchone()[0]
    assert tuple(issue) == ("pending_correction", None)
    assert correction_events == 0
    assert review_count == 0
    assert return_count == 0


def test_import_correction_return_warns_when_high_risk_lacks_note(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)

    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    _set_correction_cells(ws, 2, result="已修复")
    wb.save(paths[0])

    result = import_correction_return(app_config, paths[0])

    assert result.matched_count == 1
    assert result.review_warnings == [f"第2行高风险问题缺少整改说明：{ws['A2'].value}"]
    assert result.auto_review["still_invalid"] == 1
    with connect(app_config) as conn:
        row = conn.execute("select warning_count, warnings_json from correction_returns").fetchone()
        assert row["warning_count"] == 1
        assert "高风险问题缺少整改说明" in row["warnings_json"]
        status = conn.execute("select status from issues where issue_code = ?", (ws["A2"].value,)).fetchone()["status"]
        assert status == "still_invalid"


def test_import_correction_return_auto_closes_not_required_with_note(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)

    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    _set_correction_cells(ws, 2, result="无需整改", note="经核实该站点执行特殊电价备案")
    wb.save(paths[0])

    result = import_correction_return(app_config, paths[0])

    assert result.matched_count == 1
    assert result.auto_review["not_required"] == 1
    with connect(app_config) as conn:
        status = conn.execute("select status from issues where issue_code = ?", (ws["A2"].value,)).fetchone()["status"]
        assert status == "not_required"


def test_import_correction_return_reports_duplicate_issue_codes(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)

    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    _set_correction_cells(ws, 2, result="已修复", note="已补正")
    duplicate = [cell.value for cell in ws[2]]
    headers = {cell.value: index for index, cell in enumerate(ws[1])}
    duplicate[headers["整改结果"]] = "已修复"
    duplicate[headers["整改说明"]] = "重复回传"
    ws.append(duplicate)
    duplicate_row_number = ws.max_row
    wb.save(paths[0])

    result = import_correction_return(app_config, paths[0])

    assert result.matched_count == 1
    assert result.errors == [f"第{duplicate_row_number}行问题编号重复：{ws['A2'].value}"]


def test_import_correction_return_skips_blank_rows(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)

    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    _set_correction_cells(ws, 2, result="已修复", note="已补正")
    ws.append([""] * 15)
    wb.save(paths[0])

    result = import_correction_return(app_config, paths[0])

    assert result.matched_count == 1
    assert result.errors == []


def test_import_correction_return_skips_issue_code_without_correction(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)

    result = import_correction_return(app_config, paths[0])

    assert result.matched_count == 0
    assert result.errors == []
    with connect(app_config) as conn:
        status = conn.execute("select status from issues limit 1").fetchone()["status"]
        assert status == "pending_correction"


def test_import_correction_return_reports_missing_issue_sheet(app_config, tmp_path):
    initialize_database(app_config)
    path = tmp_path / "malformed_return.xlsx"
    wb = Workbook()
    wb.active.title = "其他"
    wb.save(path)

    result = import_correction_return(app_config, path)

    assert result.matched_count == 0
    assert result.errors == ["缺少 sheet：整改问题清单"]
    with connect(app_config) as conn:
        row = conn.execute("select error_count, errors_json from correction_returns").fetchone()
        assert row["error_count"] == 1


def test_import_correction_return_rejects_archived_batch(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    paths = export_city_issue_packages(app_config, imported.batch_id)
    with connect(app_config) as conn:
        conn.execute(
            "update issues set status = 'closed' where batch_id = ?",
            (imported.batch_id,),
        )
        conn.execute("update import_batches set status = 'returning' where id = ?", (imported.batch_id,))
    archive_batch(app_config, imported.batch_id)

    wb = load_workbook(paths[0])
    ws = wb["整改问题清单"]
    _set_correction_cells(ws, 2, result="已修复")
    wb.save(paths[0])

    with pytest.raises(ValueError, match="batch is archived"):
        import_correction_return(app_config, paths[0])


def _specialist_return_workbook(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        raw = conn.execute(
            """
            select id, row_json
              from raw_rows
             where batch_id = ? and ledger_type = 'electricity'
            """,
            (imported.batch_id,),
        ).fetchone()
        row = json.loads(raw["row_json"])
        row.update(
            {
                "电费单价": 1.2,
                "用电量": 100,
                "电费金额": 300,
                "供电方式": "转供电",
                "转供电合同情况": "无",
            }
        )
        conn.execute(
            "update raw_rows set row_json = ? where id = ?",
            (json.dumps(row, ensure_ascii=False), raw["id"]),
        )
    run_audit(app_config, imported.batch_id)
    run_electricity_analysis(app_config, imported.batch_id)
    path = export_electricity_opportunities(app_config, imported.batch_id)
    export_city_issue_packages(app_config, imported.batch_id)
    wb = load_workbook(path)
    ws = wb["整改问题清单"]
    assert ws.max_row >= 2
    headers = {cell.value: cell.column for cell in ws[1]}
    issue_code = ws.cell(row=2, column=headers["问题编号"]).value
    opportunity_code = ws.cell(row=2, column=headers["机会编号"]).value
    return path, issue_code, opportunity_code


def _set_correction_cells(ws, row_number: int, result: str = "", note: str = "") -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    if result:
        ws.cell(row=row_number, column=headers["整改结果"]).value = result
    if note:
        ws.cell(row=row_number, column=headers["整改说明"]).value = note


def _set_specialist_cells(
    ws,
    row_number: int,
    *,
    result: object = None,
    note: object = None,
    verified: object = None,
    realized: object = None,
) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    if result is not None:
        ws.cell(row=row_number, column=headers["整改结果"]).value = result
    if note is not None:
        ws.cell(row=row_number, column=headers["整改说明"]).value = note
    if verified is not None:
        ws.cell(row=row_number, column=headers["核实可追回金额"]).value = verified
    if realized is not None:
        ws.cell(row=row_number, column=headers["实际落实金额"]).value = realized
