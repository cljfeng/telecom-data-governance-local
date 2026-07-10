from governance_app.analytics import dashboard_summary
from openpyxl import load_workbook

from governance_app.archive import archive_batch, archive_precheck, export_notice_report
from governance_app.audit_engine import run_audit
import pytest

import governance_app.backup as backup_service
import governance_app.settings_service as settings_service
from governance_app.backup import create_backup, restore_backup
from governance_app.db import connect, initialize_database
from governance_app.exporter import export_city_issue_packages
from governance_app.importer import import_workbook
from governance_app.settings_service import restore_backup_safely


def test_dashboard_summary_counts_batch_and_issues(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)

    summary = dashboard_summary(app_config, imported.batch_id)

    assert summary["batch_id"] == imported.batch_id
    assert summary["ledger_counts"]["site"] == 1
    assert "issues_by_city" in summary
    assert summary["issues_by_rule"][0]["rule_name"]
    assert summary["issues_by_severity"][0]["severity"] == "high"
    assert summary["issues_by_ledger_type"][0]["ledger_type"] == "electricity"
    assert summary["issue_categories"][0]["rule_name"]
    assert summary["city_rule_matrix"][0]["city"] == "杭州"
    assert {row["rule_name"] for row in summary["city_rule_matrix"]} >= {"电费高单价", "电价异常"}
    assert summary["city_ledger_matrix"][0]["ledger_type"] == "electricity"
    assert summary["rule_effectiveness"][0]["rule_name"]
    assert summary["rule_effectiveness"][0]["confidence_label"] == "确定性问题"
    assert summary["rule_effectiveness"][0]["open_count"] >= 1
    assert summary["open_issue_count"] >= 1
    assert summary["closure_rate"] == 0.0


def test_dashboard_summary_normalizes_city_aliases(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            """
            update ledger_rows
               set city = '兰州市'
             where ledger_type = 'electricity'
            """
        )
        conn.execute("update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'")
    run_audit(app_config, imported.batch_id)
    with connect(app_config) as conn:
        conn.execute("update issues set city = '兰州' where id in (select id from issues limit 1)")
        conn.execute(
            """
            insert into issues(issue_code, audit_result_id, batch_id, city, district, ledger_type, rule_id, severity, message, suggestion)
            select 'ISS-extra-city', audit_result_id, batch_id, '兰州市', district, ledger_type, rule_id, severity, message, suggestion
              from issues
             limit 1
            """
        )

    summary = dashboard_summary(app_config, imported.batch_id)

    assert summary["issues_by_city"] == [{"city": "兰州", "count": 3}]


def test_export_notice_report_writes_issue_statistics_workbook(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)

    path = export_notice_report(app_config, imported.batch_id)

    wb = load_workbook(path)
    assert "通报总览" in wb.sheetnames
    assert "地市问题统计" in wb.sheetnames
    assert "分类统计" in wb.sheetnames
    assert "问题明细" in wb.sheetnames
    assert wb["分类统计"]["A1"].value == "分类维度"
    assert wb["问题明细"]["A2"].value


def test_backup_and_restore_database(app_config, sample_workbook):
    initialize_database(app_config)
    import_workbook(app_config, sample_workbook)

    backup_path = create_backup(app_config)

    app_config.database_path.unlink()
    restore_backup(app_config, backup_path)

    with connect(app_config) as conn:
        count = conn.execute("select count(*) as c from import_batches").fetchone()["c"]
        assert count == 1


def test_create_backup_does_not_overwrite_same_second_backup(app_config, sample_workbook):
    initialize_database(app_config)
    import_workbook(app_config, sample_workbook)

    first = create_backup(app_config)
    second = create_backup(app_config)

    assert first != second
    assert first.exists()
    assert second.exists()


def test_create_backup_passes_integrity_check(app_config, sample_workbook):
    initialize_database(app_config)
    import_workbook(app_config, sample_workbook)

    backup_path = create_backup(app_config)

    backup_service.check_database_integrity(backup_path)


def test_safe_restore_rejects_corrupt_backup_without_changing_active_database(app_config, sample_workbook):
    initialize_database(app_config)
    import_workbook(app_config, sample_workbook)
    backup_dir = app_config.workspace_dir / "backups"
    backup_dir.mkdir()
    corrupt = backup_dir / "corrupt.sqlite3"
    corrupt.write_bytes(b"not sqlite")

    with pytest.raises(ValueError, match="完整性校验失败"):
        restore_backup_safely(app_config, corrupt)

    with connect(app_config) as conn:
        count = conn.execute("select count(*) as c from import_batches").fetchone()["c"]
    assert count == 1


def test_safe_restore_rolls_back_when_post_restore_initialization_fails(app_config, sample_workbook, monkeypatch):
    initialize_database(app_config)
    import_workbook(app_config, sample_workbook)
    source_backup = create_backup(app_config)
    import_workbook(app_config, sample_workbook)

    monkeypatch.setattr(
        settings_service,
        "initialize_database",
        lambda _config: (_ for _ in ()).throw(RuntimeError("post restore failed")),
    )

    with pytest.raises(ValueError, match="已还原恢复前数据库"):
        restore_backup_safely(app_config, source_backup)

    with connect(app_config) as conn:
        count = conn.execute("select count(*) as c from import_batches").fetchone()["c"]
    assert count == 2


def test_restore_backup_rejects_file_outside_backup_dir(app_config, sample_workbook):
    initialize_database(app_config)
    import_workbook(app_config, sample_workbook)
    outside = app_config.workspace_dir / "outside.sqlite3"
    outside.write_text("not a backup", encoding="utf-8")

    try:
        restore_backup(app_config, outside)
    except ValueError as exc:
        assert "backups" in str(exc)
    else:
        raise AssertionError("restore_backup should reject files outside backups directory")


def test_archive_batch_requires_returning_or_reviewed_batch(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    run_audit(app_config, imported.batch_id)

    try:
        archive_batch(app_config, imported.batch_id)
    except ValueError as exc:
        assert "batch must be ready for archive" in str(exc)
    else:
        raise AssertionError("archive_batch should reject batches before correction return")


def test_archive_batch_writes_operation_log_sheet_and_locks_batch(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    export_city_issue_packages(app_config, imported.batch_id)
    with connect(app_config) as conn:
        conn.execute("update issues set status = 'closed' where batch_id = ?", (imported.batch_id,))
        conn.execute("update import_batches set status = 'returning' where id = ?", (imported.batch_id,))

    path = archive_batch(app_config, imported.batch_id)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "操作日志" in wb.sheetnames
    assert "版本与规则快照" in wb.sheetnames
    assert wb["版本与规则快照"]["A1"].value == "项目"
    issue_ws = wb["问题清单"]
    assert issue_ws["G1"].value == "规则分类"
    assert issue_ws["G2"].value == "问题稽核"
    assert issue_ws["I1"].value == "规则名称"
    assert issue_ws["I2"].value == "电费高单价"
    with connect(app_config) as conn:
        batch = conn.execute("select status, is_archived from import_batches where id = ?", (imported.batch_id,)).fetchone()
        assert batch["status"] == "archived"
        assert batch["is_archived"] == 1


def test_archive_precheck_reports_open_issues_before_archive(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    export_city_issue_packages(app_config, imported.batch_id)
    with connect(app_config) as conn:
        conn.execute("update import_batches set status = 'returning' where id = ?", (imported.batch_id,))

    result = archive_precheck(app_config, imported.batch_id)

    assert result["ready"] is False
    assert result["open_issue_count"] == 2
    assert result["status_counts"]["pending_correction"] == 2
    assert result["blockers"][0]["type"] == "open_issues"
    assert any(item["type"] == "high_risk_open" for item in result["risk_items"])


def test_archive_batch_adds_risk_rule_and_open_issue_sheets(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    run_audit(app_config, imported.batch_id)
    export_city_issue_packages(app_config, imported.batch_id)
    with connect(app_config) as conn:
        conn.execute("update issues set status = 'closed' where batch_id = ?", (imported.batch_id,))
        conn.execute("update import_batches set status = 'returning' where id = ?", (imported.batch_id,))

    path = archive_batch(app_config, imported.batch_id)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    overview = wb["归档总览"]
    overview_values = {row[0].value: row[1].value for row in overview.iter_rows(min_row=1, max_col=2)}
    assert overview_values["闭环率"] == 100.0
    assert overview_values["待复核问题数"] == 0
    assert "规则命中排行" in wb.sheetnames
    assert "风险等级分布" in wb.sheetnames
    assert "未闭环问题" in wb.sheetnames
    assert "专项复盘" in wb.sheetnames
    assert wb["规则命中排行"]["A1"].value == "规则分类"
    assert wb["规则命中排行"]["C2"].value == "电费高单价"
    assert wb["风险等级分布"]["A2"].value == "高"
    assert wb["未闭环问题"]["A2"].value is None
    assert wb["专项复盘"]["A1"].value == "复盘项"
    assert wb["专项复盘"]["A2"].value == "闭环率"
