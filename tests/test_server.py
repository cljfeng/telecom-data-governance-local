import json
from io import BytesIO
from pathlib import Path

import pytest

import governance_app.server as server_module
from governance_app.audit_engine import run_audit
from governance_app.config import AppConfig
from governance_app.db import SCHEMA_VERSION, connect, initialize_database
from governance_app.importer import import_workbook
from governance_app.operation_guard import exclusive_operation
from governance_app.server import create_app
from governance_app.workflow import list_issues, update_issue_status


def _multipart_upload_body(
    file_path: Path,
    fields: dict[str, str] | None = None,
    file_field: str = "file",
) -> tuple[str, bytes]:
    boundary = "----codex-test-boundary"
    parts: list[bytes] = []
    for name, value in (fields or {}).items():
        parts.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")
        )
    parts.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="{file_field}"; filename="{file_path.name}"\r\n'
            "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8")
        + file_path.read_bytes()
        + b"\r\n"
    )
    parts.append(f"--{boundary}--\r\n".encode("utf-8"))
    return f"multipart/form-data; boundary={boundary}", b"".join(parts)


def test_health_endpoint_returns_ok(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("GET", "/api/health")

    assert status == 200
    assert json.loads(body)["status"] == "ok"


def test_version_endpoint_returns_runtime_and_schema_versions(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("GET", "/api/version")

    payload = json.loads(body)
    assert status == 200
    assert headers["content-type"] == "application/json; charset=utf-8"
    assert payload["app_version"] == "0.2.0"
    assert payload["template_version"] == "2026-05-05"
    assert payload["schema_version"] == SCHEMA_VERSION
    assert payload["python_version"]


def test_dashboard_endpoint_returns_json(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("GET", "/api/dashboard?batch_id=1")

    assert status == 200
    assert headers["content-type"] == "application/json; charset=utf-8"
    assert "ledger_counts" in json.loads(body)


def test_dashboard_endpoint_rejects_invalid_batch_id(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("GET", "/api/dashboard?batch_id=abc")

    assert status == 400
    assert headers["content-type"] == "application/json; charset=utf-8"
    assert json.loads(body)["error"] == "invalid batch_id"


def test_import_audit_export_and_correction_endpoints(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/import",
        json.dumps({"path": str(sample_workbook)}),
    )

    assert status == 200
    assert json.loads(body)["batch_id"] == 1

    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/audit",
        json.dumps({"batch_id": 1}),
    )

    assert status == 200
    assert json.loads(body)["audit_run_id"] == 1

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/export",
        json.dumps({"batch_id": 1}),
    )

    assert status == 200
    exported = json.loads(body)["paths"]
    assert exported

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/corrections",
        json.dumps({"path": exported[0]}),
    )

    assert status == 200
    assert "matched_count" in json.loads(body)


def test_heavy_operation_conflict_returns_409(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    app = create_app(app_config)
    with exclusive_operation(app_config, "test"):
        status, _headers, body = app.handle_test_request(
            "POST", "/api/audit", json.dumps({"batch_id": imported.batch_id})
        )

    assert status == 409
    assert json.loads(body)["error"] == "系统正在执行其他数据操作，请稍后重试"


def test_operation_lock_releases_after_exception(app_config):
    with pytest.raises(RuntimeError, match="boom"):
        with exclusive_operation(app_config, "failing"):
            raise RuntimeError("boom")

    with exclusive_operation(app_config, "next"):
        pass


@pytest.mark.parametrize("value", [None, "invalid", "-1"])
def test_content_length_rejects_missing_invalid_and_negative_values(value):
    parser = getattr(server_module, "_content_length", None)
    assert parser is not None

    length, error = parser(value)

    assert length is None
    assert error[0] == 400


def test_content_length_rejects_request_above_100_mib():
    parser = getattr(server_module, "_content_length", None)
    maximum = getattr(server_module, "MAX_REQUEST_BODY_BYTES", None)
    assert parser is not None
    assert maximum == 100 * 1024 * 1024

    length, error = parser(str(maximum + 1))

    assert length is None
    assert error[0] == 413
    assert "100 MiB" in json.loads(error[2])["error"]


def test_oversized_request_is_rejected_without_reading_body(app_config):
    maximum = getattr(server_module, "MAX_REQUEST_BODY_BYTES", None)
    assert maximum is not None

    class NoRead(BytesIO):
        def read(self, *args, **kwargs):
            raise AssertionError("request body must not be read")

    handler = object.__new__(server_module.RequestHandler)
    handler.config = app_config
    handler.path = "/api/import/upload"
    handler.headers = {"content-length": str(maximum + 1), "content-type": "multipart/form-data"}
    handler.rfile = NoRead()
    handler.wfile = BytesIO()
    statuses = []
    handler.send_response = statuses.append
    handler.send_header = lambda _key, _value: None
    handler.end_headers = lambda: None

    handler.do_POST()

    assert statuses == [413]


def test_export_endpoint_accepts_province_mode(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    app.handle_test_request("POST", "/api/audit", json.dumps({"batch_id": 1}))

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/export",
        json.dumps({"batch_id": 1, "mode": "province"}),
    )

    assert status == 200
    paths = json.loads(body)["paths"]
    assert len(paths) == 1
    assert "全省" in paths[0]


def test_notice_report_endpoint_exports_excel(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    with connect(app_config) as conn:
        conn.execute(
            "update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'"
        )
    app.handle_test_request("POST", "/api/audit", json.dumps({"batch_id": 1}))

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/reports/notice",
        json.dumps({"batch_id": 1}),
    )

    assert status == 200
    assert json.loads(body)["path"].endswith(".xlsx")


def test_reset_endpoint_requires_confirmation(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/reset",
        json.dumps({"confirmation": "reset"}),
    )

    assert status == 400
    assert "复位" in json.loads(body)["error"]

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/reset",
        json.dumps({"confirmation": "复位"}),
    )

    assert status == 200
    assert json.loads(body)["cleared"] is True
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 0


def test_import_preview_endpoint_returns_counts_without_creating_batch(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/import/preview",
        json.dumps({"path": str(sample_workbook)}),
    )

    data = json.loads(body)
    assert status == 200
    assert data["ok"] is True
    assert data["batch_name"] == "sample_template"
    assert data["ledger_counts"]["site"] == 1
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 0


def test_import_preview_upload_endpoint_accepts_selected_file(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    content_type, body = _multipart_upload_body(sample_workbook)

    status, headers, response_body = app.handle_test_upload_request("/api/import/preview/upload", content_type, body)

    data = json.loads(response_body)
    assert status == 200
    assert data["ok"] is True
    assert data["ledger_counts"]["site"] == 1
    assert (app_config.data_dir / "uploads").exists()
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from import_batches").fetchone()["c"] == 0


def test_rules_endpoint_includes_tuning_recommendations_for_current_batch(app_config, sample_workbook):
    initialize_database(app_config)
    imported = import_workbook(app_config, sample_workbook)
    with connect(app_config) as conn:
        conn.execute("update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'")
    run_audit(app_config, imported.batch_id)
    issue = next(issue for issue in list_issues(app_config, imported.batch_id, {}) if issue["rule_id"] == "electricity_price_range")
    update_issue_status(app_config, issue["issue_code"], "not_required")
    app = create_app(app_config)

    status, _headers, body = app.handle_test_request("GET", f"/api/rules?batch_id={imported.batch_id}")

    assert status == 200
    rule = next(item for item in json.loads(body)["rules"] if item["rule_id"] == "electricity_price_range")
    assert rule["effectiveness"]["total_count"] == 1
    assert rule["effectiveness"]["not_required_rate"] == 100.0
    assert rule["tuning_recommendation"]["level"] == "warning"
    assert "无需整改率较高" in rule["tuning_recommendation"]["message"]


def test_electricity_analysis_endpoints(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    with connect(app_config) as conn:
        raw = conn.execute("select id, row_json from raw_rows where ledger_type = 'electricity'").fetchone()
        row = json.loads(raw["row_json"])
        row.update({"电费单价": 1.2, "用电量": 100, "电费金额": 300, "供电方式": "转供电", "转供电合同情况": "无"})
        conn.execute("update raw_rows set row_json = ? where id = ?", (json.dumps(row, ensure_ascii=False), raw["id"]))
    app.handle_test_request("POST", "/api/audit", json.dumps({"batch_id": 1}))

    status, _headers, body = app.handle_test_request("POST", "/api/batches/1/electricity-analysis/run")
    assert status == 200
    assert json.loads(body)["opportunity_count"] >= 3

    status, _headers, body = app.handle_test_request("GET", "/api/batches/1/electricity-analysis/summary")
    assert status == 200
    assert json.loads(body)["batch_id"] == 1

    status, _headers, body = app.handle_test_request("GET", "/api/batches/1/electricity-analysis/opportunities?limit=2&offset=0")
    assert status == 200
    page = json.loads(body)
    assert len(page["opportunities"]) == 2
    assert page["total"] >= 3
    assert page["limit"] == 2
    assert page["offset"] == 0

    status, _headers, body = app.handle_test_request("POST", "/api/batches/1/electricity-analysis/export")
    assert status == 200
    assert json.loads(body)["path"].endswith(".xlsx")


def test_electricity_analysis_rejects_invalid_batch_path(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, _headers, body = app.handle_test_request("POST", "/api/batches/not-a-number/electricity-analysis/run")

    assert status == 400
    assert json.loads(body)["error"] == "invalid batch_id"


def test_tower_rent_analysis_rejects_invalid_batch_path(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, _headers, body = app.handle_test_request("POST", "/api/batches/not-a-number/tower-rent-analysis/run")

    assert status == 400
    assert json.loads(body)["error"] == "invalid batch_id"


def test_tower_rent_analysis_endpoint_reports_missing_batch(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, _headers, body = app.handle_test_request("GET", "/api/batches/999/tower-rent-analysis/summary")

    assert status == 400
    assert json.loads(body)["error"] == "批次不存在"


def test_import_upload_endpoint_imports_selected_file(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    content_type, body = _multipart_upload_body(sample_workbook, {"strategy": "new"})

    status, headers, response_body = app.handle_test_upload_request("/api/import/upload", content_type, body)

    data = json.loads(response_body)
    assert status == 200
    assert data["batch_id"] == 1
    assert data["ledger_counts"]["site"] == 1
    with connect(app_config) as conn:
        batch = conn.execute("select source_file, name from import_batches where id = 1").fetchone()
        assert "/uploads/" in batch["source_file"]
        assert batch["name"] == "sample_template"


def test_import_upload_endpoint_reports_missing_file_without_generic_http_400(app_config):
    initialize_database(app_config)
    app = create_app(app_config)
    boundary = "----codex-test-boundary"
    content_type = f"multipart/form-data; boundary={boundary}"
    body = (
        f"--{boundary}\r\n"
        'Content-Disposition: form-data; name="strategy"\r\n\r\n'
        "new\r\n"
        f"--{boundary}--\r\n"
    ).encode("utf-8")

    status, headers, response_body = app.handle_test_upload_request("/api/import/upload", content_type, body)

    assert status == 400
    assert json.loads(response_body)["error"] == "请选择台账文件"


def test_maintenance_compact_endpoint_removes_upload_cache(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    content_type, body = _multipart_upload_body(sample_workbook, {"strategy": "new"})
    app.handle_test_upload_request("/api/import/upload", content_type, body)

    status, headers, response_body = app.handle_test_request(
        "POST",
        "/api/maintenance/compact",
        json.dumps({"clear_uploads": True}),
    )

    data = json.loads(response_body)
    assert status == 200
    assert data["removed_uploads"] == 1
    assert "after_bytes" in data


def test_import_endpoint_accepts_append_and_replace_strategy(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/import",
        json.dumps({"path": str(sample_workbook), "strategy": "append", "batch_id": batch_id}),
    )

    assert status == 200
    assert json.loads(body)["batch_id"] == batch_id
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from ledger_rows where batch_id = ?", (batch_id,)).fetchone()["c"] == 8

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/import",
        json.dumps({"path": str(sample_workbook), "strategy": "replace", "batch_id": batch_id}),
    )

    assert status == 200
    assert json.loads(body)["batch_id"] == batch_id
    with connect(app_config) as conn:
        assert conn.execute("select count(*) as c from ledger_rows where batch_id = ?", (batch_id,)).fetchone()["c"] == 4


def test_import_recent_files_and_error_export_endpoints(app_config, workbook_missing_site_code):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/import/preview",
        json.dumps({"path": str(workbook_missing_site_code)}),
    )

    assert status == 400
    preview = json.loads(body)
    assert preview["error_export_path"].endswith(".xlsx")

    status, headers, body = app.handle_test_request("GET", "/api/import/recent")

    assert status == 200
    recent = json.loads(body)["files"]
    assert recent[0]["path"] == str(workbook_missing_site_code)
    assert recent[0]["ok"] is False


def test_import_endpoint_returns_readable_error_with_validation_details(app_config, workbook_missing_site_code):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/import",
        json.dumps({"path": str(workbook_missing_site_code)}),
    )

    data = json.loads(body)
    assert status == 400
    assert data["error"] == "导入未通过，请按错误明细修正后重试"
    assert data["errors"][0]["field_name"] == "电信站址编码"


def test_workbench_management_endpoints(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/batches",
        json.dumps({"name": "2026年专项治理"}),
    )

    assert status == 200
    batch_id = json.loads(body)["batch_id"]

    status, headers, body = app.handle_test_request("GET", "/api/batches")

    assert status == 200
    assert json.loads(body)["batches"][0]["name"] == "2026年专项治理"

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/batches/current",
        json.dumps({"batch_id": batch_id}),
    )

    assert status == 200
    assert json.loads(body)["status"] == "selected"

    status, headers, body = app.handle_test_request("GET", f"/api/workflow?batch_id={batch_id}")

    assert status == 200
    assert json.loads(body)["next_action"] == "导入台账"


def test_issue_city_progress_status_and_archive_endpoints(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]
    with connect(app_config) as conn:
        conn.execute("update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'")
    app.handle_test_request("POST", "/api/audit", json.dumps({"batch_id": batch_id}))
    status, headers, body = app.handle_test_request("POST", "/api/export", json.dumps({"batch_id": batch_id}))
    exported = json.loads(body)["paths"][0]

    status, headers, body = app.handle_test_request("GET", f"/api/issues?batch_id={batch_id}&city=杭州")

    assert status == 200
    payload = json.loads(body)
    issue = payload["issues"][0]
    assert issue["city"] == "杭州"
    matching_rule = next(rule for rule in payload["rules"] if rule["rule_id"] == issue["rule_id"])
    assert matching_rule["issue_count"] >= 1

    status, headers, body = app.handle_test_request("GET", f"/api/issues?batch_id={batch_id}&rule_id={issue['rule_id']}")

    assert status == 200
    assert json.loads(body)["issues"][0]["rule_id"] == issue["rule_id"]

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/issues/status",
        json.dumps({"issue_code": issue["issue_code"], "status": "closed"}),
    )

    assert status == 200
    assert json.loads(body)["status"] == "updated"

    status, headers, body = app.handle_test_request("GET", f"/api/city-progress?batch_id={batch_id}")

    assert status == 200
    assert json.loads(body)["cities"][0]["completion_rate"] == 50.0

    from openpyxl import load_workbook

    wb = load_workbook(exported)
    ws = wb["整改问题清单"]
    headers_by_name = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers_by_name["整改结果"]).value = "已修复"
    ws.cell(row=2, column=headers_by_name["整改说明"]).value = "已补正"
    wb.save(exported)

    status, headers, body = app.handle_test_request("POST", "/api/corrections", json.dumps({"path": exported}))

    assert status == 200

    status, headers, body = app.handle_test_request("GET", f"/api/issues?batch_id={batch_id}")
    for current_issue in json.loads(body)["issues"]:
        status, headers, body = app.handle_test_request(
            "POST",
            "/api/issues/status",
            json.dumps(
                {"issue_code": current_issue["issue_code"], "status": "closed"}
            ),
        )
        assert status == 200

    status, headers, body = app.handle_test_request("POST", "/api/archive", json.dumps({"batch_id": batch_id}))

    assert status == 200
    assert json.loads(body)["path"].endswith("专项治理归档汇总.xlsx")


def test_export_endpoint_returns_json_error_before_audit(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]

    status, headers, body = app.handle_test_request("POST", "/api/export", json.dumps({"batch_id": batch_id}))

    assert status == 400
    assert json.loads(body)["error"] == "batch must be audited before export"


def test_correction_upload_endpoint_imports_selected_file(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]
    with connect(app_config) as conn:
        conn.execute("update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'")
    app.handle_test_request("POST", "/api/audit", json.dumps({"batch_id": batch_id}))
    status, headers, body = app.handle_test_request("POST", "/api/export", json.dumps({"batch_id": batch_id}))
    exported = Path(json.loads(body)["paths"][0])

    from openpyxl import load_workbook

    wb = load_workbook(exported)
    ws = wb["整改问题清单"]
    headers_by_name = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row=2, column=headers_by_name["整改结果"]).value = "已修复"
    ws.cell(row=2, column=headers_by_name["整改说明"]).value = "已按原台账修正"
    wb.save(exported)
    content_type, upload_body = _multipart_upload_body(exported)

    status, headers, body = app.handle_test_upload_request("/api/corrections/upload", content_type, upload_body)

    assert status == 200
    assert json.loads(body)["matched_count"] == 1


def test_archive_endpoint_returns_json_error_before_return(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]

    status, headers, body = app.handle_test_request("POST", "/api/archive", json.dumps({"batch_id": batch_id}))

    assert status == 400
    assert json.loads(body)["error"] == "batch must be ready for archive"


def test_archive_precheck_endpoint_reports_blockers(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]
    with connect(app_config) as conn:
        conn.execute("update raw_rows set row_json = replace(row_json, '0.8', '9.9') where ledger_type = 'electricity'")
    app.handle_test_request("POST", "/api/audit", json.dumps({"batch_id": batch_id}))
    app.handle_test_request("POST", "/api/export", json.dumps({"batch_id": batch_id}))
    with connect(app_config) as conn:
        conn.execute("update import_batches set status = 'returning' where id = ?", (batch_id,))

    status, headers, body = app.handle_test_request("GET", f"/api/archive/precheck?batch_id={batch_id}")

    assert status == 200
    payload = json.loads(body)
    assert payload["ready"] is False
    assert payload["open_issue_count"] == 2

    status, headers, body = app.handle_test_request(
        "POST", "/api/archive", json.dumps({"batch_id": batch_id})
    )

    assert status == 400
    assert json.loads(body)["error"] == "batch must be ready for archive"


def test_rules_api_lists_settings_and_updates_rule_config(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("GET", "/api/rules")

    assert status == 200
    payload = json.loads(body)
    price_rule = next(rule for rule in payload["rules"] if rule["rule_id"] == "electricity_price_range")
    assert price_rule["enabled"] is True
    assert price_rule["config"] == {}
    assert price_rule["category"] == "problem_audit"
    assert price_rule["parameters"][0]["key"] == "max"
    quality_rule = next(rule for rule in payload["rules"] if rule["rule_id"] == "site_code_missing_in_master")
    assert quality_rule["category"] == "data_quality"
    reading_rule = next(rule for rule in payload["rules"] if rule["rule_id"] == "electricity_meter_reading_reverse")
    assert reading_rule["category"] == "data_quality"
    amount_rule = next(rule for rule in payload["rules"] if rule["rule_id"] == "electricity_amount_calculation_mismatch")
    assert [param["key"] for param in amount_rule["parameters"]] == ["variance_ratio", "variance_min"]

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/rules/settings",
        json.dumps({"rule_id": "electricity_price_range", "enabled": False, "config": {"max": 3}}, ensure_ascii=False),
    )

    assert status == 200
    assert json.loads(body) == {"status": "updated"}
    status, headers, body = app.handle_test_request("GET", "/api/rules")
    payload = json.loads(body)
    price_rule = next(rule for rule in payload["rules"] if rule["rule_id"] == "electricity_price_range")
    assert price_rule["enabled"] is False
    assert price_rule["config"] == {"max": 3}


def test_ledger_rows_endpoint_filters_current_batch_data(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    status, headers, body = app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    batch_id = json.loads(body)["batch_id"]

    status, headers, body = app.handle_test_request(
        "GET",
        f"/api/ledger-rows?batch_id={batch_id}&ledger_type=electricity&city=杭州&site_code=HZ001",
    )

    assert status == 200
    rows = json.loads(body)["rows"]
    assert len(rows) == 1
    assert rows[0]["field_groups"]["电表报账"]["电表户号"] == "M001"


def test_backup_and_restore_endpoints(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("POST", "/api/backup", "{}")

    assert status == 200
    backup_path = json.loads(body)["path"]

    status, headers, body = app.handle_test_request(
        "POST",
        "/api/restore",
        json.dumps({"path": backup_path}),
    )

    assert status == 200
    assert json.loads(body)["status"] == "restored"


def test_settings_endpoint_reports_local_paths(app_config):
    initialize_database(app_config)
    app = create_app(app_config)

    status, headers, body = app.handle_test_request("GET", "/api/settings")

    assert status == 200
    payload = json.loads(body)
    assert payload["database_path"].endswith("governance.sqlite3")
    assert payload["export_dir"].endswith("exports")
    assert payload["backup_dir"].endswith("backups")
    assert payload["template_version"] == "2026-05-05"
    version_status, _version_headers, version_body = app.handle_test_request("GET", "/api/version")
    assert version_status == 200
    assert payload["template_version"] == json.loads(version_body)["template_version"]


def test_restore_endpoint_creates_safety_backup(app_config, sample_workbook):
    initialize_database(app_config)
    app = create_app(app_config)
    app.handle_test_request("POST", "/api/import", json.dumps({"path": str(sample_workbook)}))
    status, headers, body = app.handle_test_request("POST", "/api/backup", "{}")
    backup_path = json.loads(body)["path"]

    status, headers, body = app.handle_test_request("POST", "/api/restore", json.dumps({"path": backup_path}))

    assert status == 200
    payload = json.loads(body)
    assert payload["status"] == "restored"
    assert payload["safety_backup_path"].endswith(".sqlite3")
    assert Path(payload["safety_backup_path"]).exists()


def test_static_handler_disables_browser_cache():
    from governance_app.server import RequestHandler

    assert RequestHandler.extra_static_headers()["Cache-Control"] == "no-store, max-age=0"
    assert RequestHandler.extra_static_headers()["Pragma"] == "no-cache"


def test_workbench_static_assets_use_lightweight_modules():
    static_dir = AppConfig.for_workspace(Path(".")).static_dir
    index_html = (static_dir / "index.html").read_text(encoding="utf-8")
    app_js = (static_dir / "app.js").read_text(encoding="utf-8")
    ledger_data_js = (static_dir / "ledger-data.js").read_text(encoding="utf-8")

    assert '<script type="module" src="/app.js' in index_html
    assert 'data-view="ledgerData"' in index_html
    assert "<span>数据整理</span>" in index_html
    assert 'from "/api.js' in app_js
    assert 'from "/state.js' in app_js
    assert 'from "/ui.js' in app_js
    assert 'from "/ledger-data.js' in app_js
    assert 'from "/rules.js' in app_js
    assert 'from "/settings.js' in app_js
    assert 'from "/analytics.js' in app_js
    assert "/api/ledger-rows?" in ledger_data_js
    assert 'data-view="rules"' in index_html
    assert "<span>规则设置</span>" in index_html
    assert 'data-view="settings"' in index_html
    assert "<span>本地设置</span>" in index_html
    assert (static_dir / "api.js").exists()
    assert (static_dir / "state.js").exists()
    assert (static_dir / "ui.js").exists()
    assert (static_dir / "ledger-data.js").exists()
    assert (static_dir / "rules.js").exists()
    assert (static_dir / "settings.js").exists()
    assert (static_dir / "analytics.js").exists()


def test_rules_static_module_calls_rules_api():
    static_dir = AppConfig.for_workspace(Path(".")).static_dir
    rules_js = (static_dir / "rules.js").read_text(encoding="utf-8")

    assert "/api/rules" in rules_js
    assert "/api/rules/settings" in rules_js
    assert "electricity_price_range" in rules_js
    assert "基础数据质量" in rules_js
    assert "恢复默认" in rules_js
    assert "data-rule-param" in rules_js


def test_import_page_exposes_import_strategies():
    static_dir = AppConfig.for_workspace(Path(".")).static_dir
    app_js = (static_dir / "app.js").read_text(encoding="utf-8")

    assert 'id="import-strategy"' in app_js
    assert 'value="new"' in app_js
    assert 'value="append"' in app_js
    assert 'value="replace"' in app_js


def test_operator_experience_ui_surfaces_guidance_and_review_decisions():
    static_dir = AppConfig.for_workspace(Path(".")).static_dir
    app_js = (static_dir / "app.js").read_text(encoding="utf-8")
    settings_js = (static_dir / "settings.js").read_text(encoding="utf-8")

    assert "blocked_reason" in app_js
    assert "review_suggestion" in app_js
    assert "recommended_action" in app_js
    assert "top_rules" in app_js
    assert "导出整改包会更新问题状态" in app_js
    assert "确认恢复" in settings_js
    assert "当前数据库会先自动安全备份" in settings_js


def test_settings_static_module_calls_settings_and_backup_api():
    static_dir = AppConfig.for_workspace(Path(".")).static_dir
    settings_js = (static_dir / "settings.js").read_text(encoding="utf-8")

    assert "/api/settings" in settings_js
    assert "/api/backup" in settings_js
    assert "/api/restore" in settings_js


def test_analytics_static_module_uses_dashboard_summary():
    static_dir = AppConfig.for_workspace(Path(".")).static_dir
    analytics_js = (static_dir / "analytics.js").read_text(encoding="utf-8")

    assert "/api/dashboard?batch_id=" in analytics_js
    assert "issues_by_severity" in analytics_js
    assert "closure_rate" in analytics_js
    assert "问题类型矩阵" in analytics_js
    assert "基础数据质量" in analytics_js
    assert "发电费风险" in analytics_js
    assert "导出通报" in analytics_js


def test_readme_includes_operator_flow_and_faq():
    readme = Path("README.md").read_text(encoding="utf-8")

    assert "业务人员使用流程" in readme
    assert "常见问题" in readme
    assert "验收流程" in readme


def test_start_script_documents_local_launch_command():
    script = Path("scripts/start.sh")

    assert script.exists()
    content = script.read_text(encoding="utf-8")
    assert "PYTHONPATH=src" in content
    assert "governance_app.server" in content
    assert "--port" in content


def test_check_script_runs_project_verification_commands():
    script = Path("scripts/check.sh")

    assert script.exists()
    content = script.read_text(encoding="utf-8")
    assert ".venv/bin/python -m pytest -q" in content
    assert "python -m compileall -q src" in content
    assert "src/governance_app/static/*.js" in content
    assert "node --check src/governance_app/static/app.js" not in content
    assert "bash -n scripts/start.sh" in content
    assert "bash -n scripts/build_app.sh" in content
